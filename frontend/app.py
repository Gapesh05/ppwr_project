"""Frontend Flask application for PFAS/PPWR UI."""

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session
import os
from config import Config
from io import BytesIO
import pandas as pd
import logging
from logging.handlers import RotatingFileHandler
from openpyxl.styles import PatternFill
import requests
from flask_sqlalchemy import SQLAlchemy
from models import db as models_db, Route, PFASMaterialChemicals, PFASRegulations, PFASBOMAudit, SupplierDeclarationV1, MaterialDeclarationLink, PPWRBOM, PPWRMaterialDeclarationLink, PPWRResult
from fastapi_client import (
    ingest_material_data,
    get_assessments as fastapi_get_assessments,
    get_ppwr_evaluation_summary as fastapi_get_ppwr_evaluation_summary,
    assess_with_files as fastapi_assess_with_files,
    list_supplier_declarations as fastapi_list_supplier_declarations,
    upload_supplier_declaration as fastapi_upload_supplier_declaration,
    assess_from_declaration as fastapi_assess_from_declaration,
)
import json
from datetime import datetime
import sqlalchemy
from sqlalchemy import text
from flask import Response   # if you want template download route returning CSV
from decimal import Decimal, InvalidOperation
import re
import time
from sqlalchemy import and_


# ==================== HELPERS ====================
def _build_distinct_ppwr_declarations():
    """Return one row per material from supplier_declaration_v1 table."""
    try:
        rows = SupplierDeclarationV1.query.order_by(SupplierDeclarationV1.upload_date.desc()).all()
        out = []
        for r in rows:
            out.append({
                'material_id': getattr(r, 'material_id', None),
                'original_filename': r.original_filename,
                'file_size': r.file_size,
                'upload_date': r.upload_date.isoformat() if r.upload_date else None,
            })
        return out
    except Exception:
        app.logger.exception('Failed to build distinct PPWR declarations list')
        return []

# ==================== FLASK APP CONFIG ====================

app = Flask(__name__)
app.config.from_object(Config)
app.secret_key = 'your-super-secret-key'

# Enable debug mode
app.config['DEBUG'] = True
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key')

# Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
app.logger.setLevel(logging.INFO)

# Write logs to a rotating file as well (frontend/logs/app.log)
try:
    logs_dir = os.path.join(os.path.dirname(__file__), 'logs')
    os.makedirs(logs_dir, exist_ok=True)
    file_handler = RotatingFileHandler(
        os.path.join(logs_dir, 'app.log'),
        maxBytes=2 * 1024 * 1024,  # 2 MB per file
        backupCount=5,
        encoding='utf-8'
    )
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(name)s - %(message)s'))
    app.logger.addHandler(file_handler)
    # Also attach to root logger so SQLAlchemy and werkzeug messages go to file
    logging.getLogger().addHandler(file_handler)
except Exception as _e:
    # Fallback: continue without file handler
    pass

# Database
# Use the SQLAlchemy instance defined in `models.py` so models' metadata is registered
db = models_db
db.init_app(app)
# Ensure SQLAlchemy sessions are removed at the end of each request/context
@app.teardown_appcontext
def shutdown_session(exception=None):
    try:
        db.session.remove()
    except Exception:
        pass

# PPWR: standalone evaluation page route
@app.route('/ppwr/evaluation')
def ppwr_assessment_evaluation_page():
    """
    Render the standalone PPWR assessment evaluation UI page with data from ppwr_result + ppwr_bom.
    Supports optional SKU filtering via ?sku=<value> query parameter.
    """
    try:
        # Get optional SKU filter from query string
        sku_filter = request.args.get('sku', None)
        
        # Build query: Join ppwr_result + ppwr_bom to get complete evaluation data
        # Columns needed: Component, Subcomponent, Material, Supplier, CAS_ID, Chemical, Concentration, Status
        query = db.session.query(
            PPWRBOM.component,
            PPWRBOM.component_description,
            PPWRBOM.subcomponent,
            PPWRBOM.subcomponent_description,
            PPWRBOM.material_id.label('material'),
            PPWRBOM.material_name,
            PPWRResult.supplier_name.label('supplier'),
            PPWRResult.cas_id,
            PPWRResult.chemical,
            PPWRResult.concentration,
            PPWRResult.status
        ).join(
            PPWRResult, PPWRBOM.material_id == PPWRResult.material_id, isouter=True
        )
        
        # Apply SKU filter if provided
        if sku_filter:
            query = query.filter(PPWRBOM.sku == sku_filter)
        
        results = query.all()
        
        # Build rows for template
        rows = []
        stats = {
            'total_files': 0,
            'files_downloaded': 0,
            'conformance': 0,
            'non_conformance': 0
        }
        
        for row in results:
            # Determine status
            status_display = row.status if row.status else 'Unknown'
            if status_display.lower() in ('compliant', 'conformance', 'compliance'):
                stats['conformance'] += 1
                status_color = 'success'
            elif status_display.lower() in ('non-compliant', 'non-conformance', 'non-compliance'):
                stats['non_conformance'] += 1
                status_color = 'danger'
            else:
                status_color = 'warning'
            
            rows.append({
                'component': row.component or 'No Data',
                'sub_component': row.subcomponent or 'No Data',
                'material_id': row.material or 'No Data',
                'material_name': row.material_name or 'No Data',
                'supplier': row.supplier or 'No Data',
                'cas_id': row.cas_id or 'No Data',
                'chemical': row.chemical or 'No Data',
                'concentration': f"{float(row.concentration):.2f} ppm" if row.concentration is not None else 'No Data',
                'status': status_display,
                'status_color': status_color
            })
        
        # Calculate file stats
        if sku_filter:
            total_materials = db.session.query(PPWRBOM.material_id).filter_by(sku=sku_filter).distinct().count()
            materials_with_results = db.session.query(PPWRBOM.material_id).filter_by(sku=sku_filter).join(
                PPWRResult, PPWRBOM.material_id == PPWRResult.material_id
            ).distinct().count()
        else:
            total_materials = db.session.query(PPWRBOM.material_id).distinct().count()
            materials_with_results = db.session.query(PPWRBOM.material_id).join(
                PPWRResult, PPWRBOM.material_id == PPWRResult.material_id
            ).distinct().count()
        
        stats['total_files'] = total_materials
        stats['files_downloaded'] = materials_with_results
        
        return render_template('ppwr_assessment_evaluation.html', stats=stats, rows=rows, sku=sku_filter)
    except Exception as e:
        app.logger.error(f"Error rendering PPWR assessment evaluation page: {e}", exc_info=True)
        return "Error loading evaluation page", 500

@app.route('/api/ppwr/evaluation/download-csv/<sku>')
def download_ppwr_evaluation_csv(sku):
    """Download PPWR evaluation results as CSV file."""
    try:
        app.logger.info(f"📥 Generating CSV for SKU: {sku}")
        
        # Reuse existing evaluation query
        query = db.session.query(
            PPWRBOM.component,
            PPWRBOM.subcomponent,
            PPWRBOM.material_id,
            PPWRBOM.material_name,
            PPWRResult.supplier_name,
            PPWRResult.cas_id,
            PPWRResult.chemical,
            PPWRResult.concentration,
            PPWRResult.status
        ).outerjoin(
            PPWRResult, PPWRBOM.material_id == PPWRResult.material_id
        ).filter(PPWRBOM.sku == sku)
        
        results = query.all()
        
        if not results:
            flash("No data to export", "warning")
            return redirect(url_for('ppwr_assessment_evaluation_page') + f'?sku={sku}')
        
        # Build CSV rows
        csv_lines = [
            "Component,Subcomponent,Material ID,Material Name,Supplier,CAS ID,Chemical,Concentration,Status"
        ]
        
        for row in results:
            # Escape commas in text fields
            component = (row.component or 'No Data').replace(',', ';')
            subcomponent = (row.subcomponent or 'No Data').replace(',', ';')
            material_id = (row.material_id or 'No Data').replace(',', ';')
            material_name = (row.material_name or 'No Data').replace(',', ';')
            supplier = (row.supplier_name or 'No Data').replace(',', ';')
            cas_id = (row.cas_id or 'No Data').replace(',', ';')
            chemical = (row.chemical or 'No Data').replace(',', ';')
            concentration = f"{float(row.concentration):.2f} ppm" if row.concentration is not None else 'No Data'
            status = row.status if row.status else 'Unknown'
            
            csv_lines.append(
                f'"{component}","{subcomponent}","{material_id}","{material_name}","{supplier}","{cas_id}","{chemical}","{concentration}","{status}"'
            )
        
        # Generate CSV content
        csv_content = "\n".join(csv_lines)
        csv_bytes = csv_content.encode('utf-8')
        
        # Create filename with timestamp
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        filename = f"PPWR_Evaluation_{sku}_{timestamp}.csv"
        
        return send_file(
            BytesIO(csv_bytes),
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        app.logger.error(f"CSV download failed: {e}", exc_info=True)
        flash("Error generating CSV", "danger")
        return redirect(url_for('ppwr_assessment_evaluation_page') + f'?sku={sku}')

# Defer table creation to guarded startup block in __main__ to avoid
# failing import when DB is temporarily unreachable.

# Directory where supplier declaration files are stored (on the frontend service filesystem)
SUPPLIER_DECLARATION_DIR = os.path.join(app.root_path, 'supplier_declarations')
os.makedirs(SUPPLIER_DECLARATION_DIR, exist_ok=True)

# Allowed file extensions for declarations
ALLOWED_DECL_EXT = {'.pdf', '.txt', '.csv', '.xls', '.xlsx', '.doc', '.docx'}

def _allowed_decl(filename: str) -> bool:
    if not filename:
        return False
    _, ext = os.path.splitext(filename.lower())
    return ext in ALLOWED_DECL_EXT


# Quiet favicon 404s to keep console clean during testing
@app.route('/favicon.ico')
def _favicon():
    try:
        static_favicon = os.path.join(app.root_path, 'static', 'favicon.ico')
        if os.path.exists(static_favicon):
            return send_file(static_favicon)
    except Exception:
        pass
    return ('', 204)


def _normalize_material_value(raw_val: object) -> str:
    """Normalize a material cell value to a stable material_id-like token.

    Handles multi-line cells like "1073B Tyvek\nA8362" by selecting the most ID-like token (usually the last short token).
    Falls back to the stripped raw value.
    """
    if raw_val is None:
        return ''
    try:
        s = str(raw_val).strip()
    except Exception:
        return ''

    if s == '' or s.lower() == 'nan':
        return ''

    # Split on newlines first, then common separators
    parts = [p.strip() for p in re.split(r'[\r\n]+', s) if p.strip()]
    if len(parts) == 0:
        return ''
    if len(parts) == 1:
        candidate = parts[0]
    else:
        # Prefer the last part if it looks ID-like (short, alphanumeric)
        last = parts[-1]
        if re.match(r'^[A-Za-z0-9_\-]{1,40}$', last):
            candidate = last
        else:
            # otherwise prefer the shortest part (likely an ID)
            candidate = min(parts, key=lambda x: len(x))

    # Further sanitize: remove surrounding parentheses and stray characters
    candidate = candidate.strip()
    candidate = re.sub(r'^[\(\[\{\s]+|[\)\]\}\s]+$', '', candidate)

    # If candidate contains a space and also contains a short token, pick that
    if ' ' in candidate:
        tokens = [t.strip() for t in re.split(r'[\s/,_]+', candidate) if t.strip()]
        for t in tokens[::-1]:
            if re.match(r'^[A-Za-z0-9_\-]{1,40}$', t):
                candidate = t
                break

    return candidate



# ==================== ROUTES ====================


@app.route('/')
def index():
    app.logger.info("Loading index page")
    try:
        # NEW: Query distinct products from ppwr_bom grouped by SKU, join with route table
        from sqlalchemy import func
        
        # Get distinct SKUs with their latest upload timestamp from ppwr_bom
        ppwr_data = db.session.query(
            PPWRBOM.sku,
            func.max(PPWRBOM.uploaded_at).label('uploaded_at')
        ).filter(PPWRBOM.sku.isnot(None)).group_by(PPWRBOM.sku).all()
        
        # Build SKU to upload timestamp map
        sku_to_uploaded = {row.sku: row.uploaded_at for row in ppwr_data}
        
        # Get routes for these SKUs
        skus = [row.sku for row in ppwr_data]
        routes = db.session.query(Route).filter(Route.sku.in_(skus)).all() if skus else []
        sku_to_route = {r.sku: r.route for r in routes}
        
        # Check if there was a recent upload to highlight (used as a fallback only)
        last_upload = session.pop('last_upload', None)  # contains {'timestamp': str, 'skus': [..]} or None
        uploaded_ts = None
        uploaded_skus = set()
        if last_upload:
            uploaded_ts = last_upload.get('timestamp')
            uploaded_skus = set(last_upload.get('skus', []))
            app.logger.info(f"Applying last_upload timestamp {uploaded_ts} to SKUs: {uploaded_skus}")

        products = []
        for row in ppwr_data:
            sku_val = row.sku
            # Prefer the persisted max(uploaded_at) from DB; if not available, use the recent session timestamp
            persisted_dt = sku_to_uploaded.get(sku_val)
            if persisted_dt:
                try:
                    uploaded_at_str = persisted_dt.strftime("%Y-%m-%d %H:%M:%S UTC")
                except Exception:
                    uploaded_at_str = str(persisted_dt)
            else:
                uploaded_at_str = uploaded_ts if sku_val in uploaded_skus else None

            # Derive product name from sku (could also add product column to ppwr_bom if needed)
            product_name = f"Product {sku_val}"

            products.append({
                'sku': sku_val,
                'product_name': product_name,
                'route': sku_to_route.get(sku_val, 'ppwr').lower(),
                'uploaded_at': uploaded_at_str
            })

        return render_template('index.html', products=products)
    except Exception as e:
        app.logger.error(f"Error loading products: {e}", exc_info=True)
        flash("Error loading products", "danger")
        return render_template('index.html', products=[])

@app.route('/start/<sku>')
def start_assessment_by_route(sku):
    """Redirect Start action based on route table value for the SKU.

    Routes to unified assessment page with appropriate tab parameter.
    Tab is determined by the 'route' column in the route table.
    """
    try:
        rec = db.session.query(Route).filter_by(sku=sku).first()
        route_val = (rec.route or 'ppwr').strip().lower() if rec else 'ppwr'
        
        # Redirect to unified assessment with tab parameter
        return redirect(url_for('assessment_page', sku=sku, tab=route_val))
    except Exception as e:
        app.logger.error(f"Start routing failed for {sku}: {e}", exc_info=True)
        flash('Failed to start assessment; defaulting to PPWR.', 'warning')
        return redirect(url_for('assessment_page', sku=sku, tab='ppwr'))

@app.route('/ppwr/declarations', methods=['GET'])
def ppwr_declarations_page():
    try:
        # Build a de-duplicated list (one row per material)
        declarations = _build_distinct_ppwr_declarations()

        # Fetch latest assessments for materials present in the declarations list
        material_ids = sorted({d.get('material_id') for d in declarations if d.get('material_id')})
        assessments_by_mat = {}
        for mid in material_ids:
            try:
                aresp = fastapi_get_assessments(mid)
                if aresp and aresp.get('success'):
                    items = aresp.get('assessments') or []
                    if items:
                        assessments_by_mat[mid] = items[0]
            except Exception:
                app.logger.exception(f"Failed to fetch assessments for {mid}")

        return render_template('ppwr_declarations.html', declarations=declarations, eval_result=None, evaluated_id=None, assessments=assessments_by_mat)
    except Exception as e:
        app.logger.error(f"PPWR declarations page error: {e}", exc_info=True)
        flash('Failed to fetch declarations', 'danger')
        return render_template('ppwr_declarations.html', declarations=[], eval_result=None, evaluated_id=None, assessments={})

@app.route('/ppwr/declarations/evaluate', methods=['POST'])
def ppwr_declarations_evaluate():
    try:
        material_id = request.form.get('material_id')
        if not material_id:
            flash('Missing material id', 'danger')
            return redirect(url_for('ppwr_declarations_page'))
        # Fetch local bytes and post to FastAPI /ppwr/assess
        decl = SupplierDeclaration.query.filter_by(material_id=str(material_id).strip()).first()
        if not decl or not getattr(decl, 'file_data', None):
            flash('Declaration not found or file bytes missing', 'danger')
            return redirect(url_for('ppwr_declarations_page'))

        fname = decl.original_filename or f"decl_{material_id}.pdf"
        content_type = 'application/pdf'
        assess_resp = fastapi_assess_with_files([str(material_id).strip()], [(fname, bytes(decl.file_data), content_type)])
        if not assess_resp or not assess_resp.get('success'):
            flash('Assessment failed', 'danger')

        # Build local declarations list (deduped per material)
        declarations = _build_distinct_ppwr_declarations()

        # Populate assessments so the status badges/inline blocks render
        material_ids = sorted({d.get('material_id') for d in declarations if d.get('material_id')})
        assessments_by_mat = {}
        for mid in material_ids:
            try:
                aresp = fastapi_get_assessments(mid)
                if aresp and aresp.get('success'):
                    items = aresp.get('assessments') or []
                    if items:
                        assessments_by_mat[mid] = items[0]
            except Exception:
                app.logger.exception(f"Failed to fetch assessments for {mid}")

        # Render inline result if any
        assessment_obj = None
        try:
            aresp = fastapi_get_assessments(str(material_id).strip())
            if aresp and aresp.get('success'):
                items = aresp.get('assessments') or []
                if items:
                    assessment_obj = items[0]
        except Exception:
            pass

        return render_template('ppwr_declarations.html', declarations=declarations, eval_result=assessment_obj, evaluated_id=material_id, assessments=assessments_by_mat)
    except Exception as e:
        app.logger.error(f"PPWR evaluate error: {e}", exc_info=True)
        flash('Error during evaluation', 'danger')
        return redirect(url_for('ppwr_declarations_page'))


@app.route('/ppwr/declarations/evaluate-all', methods=['POST'])
def ppwr_declarations_evaluate_all():
    """Run PPWR assessment for every listed declaration when explicitly requested.

    This prevents automatic runs on page load and lets the user trigger batch evaluation.
    """
    try:
        # Use local declarations list
        rows = SupplierDeclaration.query.filter_by(is_archived=False).order_by(SupplierDeclaration.upload_date.desc()).all()
        declarations = []
        for r in rows:
            declarations.append({
                'material_id': getattr(r, 'material_id', None),
                'original_filename': r.original_filename,
            })
        successes = 0
        failures = []
        for d in declarations:
            material_id = d.get('material_id')
            if not material_id:
                failures.append({'id': None, 'material_id': material_id, 'error': 'missing ids'})
                continue
            try:
                # Send local bytes to FastAPI /ppwr/assess
                rec = SupplierDeclaration.query.filter_by(material_id=str(material_id)).first()
                if not rec or not getattr(rec, 'file_data', None):
                    failures.append({'id': None, 'material_id': material_id, 'error': 'missing bytes'})
                    continue
                fname = rec.original_filename or f"decl_{material_id}.pdf"
                content_type = 'application/pdf'
                result = fastapi_assess_with_files([str(material_id).strip()], [(fname, bytes(rec.file_data), content_type)])
                if result and result.get('success'):
                    successes += 1
                else:
                    failures.append({'id': None, 'material_id': material_id, 'error': (result or {}).get('error', 'unknown')})
            except Exception as ex:
                app.logger.exception(f"Batch PPWR evaluate failed for material {material_id}")
                failures.append({'id': None, 'material_id': material_id, 'error': str(ex)})

        # Rebuild assessments map so UI shows latest results
        material_ids = sorted({d.get('material_id') for d in declarations if d.get('material_id')})
        assessments_by_mat = {}
        for mid in material_ids:
            try:
                aresp = fastapi_get_assessments(mid)
                if aresp and aresp.get('success'):
                    items = aresp.get('assessments') or []
                    if items:
                        assessments_by_mat[mid] = items[0]
            except Exception:
                app.logger.exception(f"Failed to fetch assessments for {mid}")

        if successes:
            flash(f"Ran PPWR assessment for {successes} declaration(s)", 'success')
        if failures:
            flash(f"{len(failures)} declaration(s) failed to evaluate", 'warning')

        return render_template('ppwr_declarations.html', declarations=declarations, eval_result=None, evaluated_id=None, assessments=assessments_by_mat)
    except Exception as e:
        app.logger.error(f"PPWR evaluate-all error: {e}", exc_info=True)
        flash('Error during batch evaluation', 'danger')
        return redirect(url_for('ppwr_declarations_page'))


@app.route('/api/ppwr/assessments/batch', methods=['POST'])
def api_ppwr_assessments_batch():
    """Return latest PPWR assessments for a list of material IDs.

    Body: { material_ids: ["A123", "B234", ...] }
    Response: { success: bool, assessments: { material_id: {..} } }
    """
    try:
        payload = request.get_json(silent=True) or {}
        material_ids = payload.get('material_ids') or []
        material_ids = [str(m).strip() for m in material_ids if str(m).strip()]
        if not material_ids:
            return jsonify({'success': False, 'error': 'material_ids required'}), 400

        assessments_by_mat = {}
        for mid in material_ids:
            try:
                aresp = fastapi_get_assessments(mid)
                if aresp and aresp.get('success'):
                    items = aresp.get('assessments') or []
                    if items:
                        assessments_by_mat[mid] = items[0]
            except Exception as e:
                app.logger.debug(f"Batch get assessments failed for {mid}: {e}")
                continue

        return jsonify({'success': True, 'assessments': assessments_by_mat}), 200
    except Exception as e:
        app.logger.error(f"Batch assessments endpoint failed: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/debug/ppwr/storage-index', methods=['GET'])
def api_debug_ppwr_storage_index():
    """Return stored path for declaration and Chroma index status if available.

    Params: material_id, decl_id
    """
    try:
        material_id = request.args.get('material_id') or ''
        decl_id = request.args.get('decl_id') or ''
        out = {
            'success': True,
            'material_id': material_id,
            'decl_id': decl_id,
            'frontend': {},
            'backend': {},
        }
        # Frontend storage details
        try:
            drec = None
            if material_id:
                drec = SupplierDeclaration.query.filter_by(material_id=material_id).first()
            if drec:
                out['frontend']['stored_path'] = getattr(drec, 'stored_path', None)
                out['frontend']['sku'] = getattr(drec, 'sku', None)
                out['frontend']['supplier_name'] = getattr(drec, 'supplier_name', None)
                # Count links for this material
                try:
                    cnt = db.session.query(PPWRMaterialDeclarationLink).filter_by(material_id=material_id).count()
                    out['frontend']['links_count'] = cnt
                except Exception:
                    out['frontend']['links_count'] = None
        except Exception as e:
            out['frontend']['error'] = str(e)

        # Backend declarations list and optional chroma chunks
        try:
            blist = fastapi_list_supplier_declarations(material_id=material_id)
            out['backend']['declarations'] = blist
        except Exception as e:
            out['backend']['declarations_error'] = str(e)
        # Attempt chroma chunks endpoint if available
        try:
            base = os.environ.get('FASTAPI_BASE_URL') or 'http://127.0.0.1:8000'
            import requests as _rq
            rr = _rq.get(f"{base}/ppwr/chroma/chunks", params={'material_id': material_id, 'decl_id': decl_id}, timeout=10)
            if rr.status_code == 200:
                out['backend']['chroma'] = rr.json()
            else:
                out['backend']['chroma_status'] = rr.status_code
        except Exception as e:
            out['backend']['chroma_error'] = str(e)
        return jsonify(out)
    except Exception as e:
        app.logger.exception('debug storage-index error')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/ppwr/material-details/<material_id>', methods=['GET'])
def api_ppwr_material_details(material_id):
    """Fetch supplier campaign and supplier name for material details expansion.
    
    Returns:
        JSON with success, material_id, supplier_name, supplier_campaign
    """
    try:
        app.logger.info(f"Fetching material details for: {material_id}")
        
        # Query PPWR BOM table for material details
        bom_row = db.session.query(PPWRBOM).filter_by(material_id=material_id).first()
        
        if not bom_row:
            return jsonify({
                'success': False,
                'error': f'Material {material_id} not found in PPWR BOM'
            }), 404
        
        # Check if supplier declaration exists (determines campaign status)
        decl = SupplierDeclaration.query.filter_by(material_id=material_id).first()
        has_declaration = decl is not None and not getattr(decl, 'is_archived', False)
        
        # Determine supplier campaign status
        supplier_campaign = 'Active' if has_declaration else None
        
        return jsonify({
            'success': True,
            'material_id': material_id,
            'supplier_name': bom_row.supplier_name or '\u2014',
            'supplier_campaign': supplier_campaign,
            'material_name': bom_row.material_name,  # Sent but not displayed in UI
            'has_declaration': has_declaration  # Sent but not displayed in UI
        }), 200
        
    except Exception as e:
        app.logger.error(f"Error fetching material details for {material_id}: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# Debug helper: list sample materials and a declaration id if available
@app.route('/api/debug/ppwr/list-materials', methods=['GET'])
def api_debug_ppwr_list_materials():
    try:
        # List distinct materials from BOM
        mats = []
        try:
            rows = db.session.query(PPWRBOM.material_id).distinct().all()
            mats = [r[0] for r in rows if r and r[0]]
        except Exception as e:
            app.logger.exception('Failed to list BOM materials')
        return jsonify({'success': True, 'materials': mats[:50]})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/ppwr/declarations/upload', methods=['POST'])
def ppwr_declarations_upload_proxy():
    try:
        file = request.files.get('file')
        if not file:
            flash('No file uploaded', 'danger')
            return redirect(url_for('ppwr_declarations_page'))

        fname = file.filename or 'document.bin'
        if not _allowed_decl(fname):
            flash('Unsupported file extension', 'danger')
            return redirect(url_for('ppwr_declarations_page'))

        sku = (request.form.get('sku') or '').strip()
        material = (request.form.get('material') or '').strip()
        supplier_name = (request.form.get('supplier_name') or '').strip()
        description = (request.form.get('description') or '').strip()

        # Infer material from filename prefix when missing
        if not material:
            try:
                base = os.path.splitext(fname)[0]
                tokens = [t for t in re.split(r'[\s_\-]+', base) if t]
                if tokens:
                    cand = tokens[0]
                    if re.match(r'^[A-Za-z0-9_\-]{1,40}$', cand):
                        material = cand
            except Exception:
                material = ''

        if not material:
            flash('Material could not be inferred from filename', 'danger')
            return redirect(url_for('ppwr_declarations_page'))

        # Validate BOM existence (warn but continue)
        try:
            if sku and not db.session.query(PPWRBOM).filter_by(sku=sku, material_id=material).first():
                app.logger.warning(f"PPWR proxy: no PPWR BOM match for sku={sku} material={material}; proceeding with upload")
                flash(f"Warning: No PPWR BOM row found for SKU '{sku}' and material '{material}'. Upload will proceed.", 'warning')
        except Exception:
            app.logger.exception('PPWR proxy: BOM validation error; proceeding with upload')
            flash('Warning: BOM validation error. Upload will proceed.', 'warning')

        # Prepare payload
        file_bytes = file.read()
        file_size = len(file_bytes) if file_bytes is not None else None
        _ext = os.path.splitext(fname.lower())[1]
        if _ext in ('.pdf',):
            doc_type = 'pdf'
        elif _ext in ('.docx', '.doc'):
            doc_type = 'docx'
        elif _ext in ('.xlsx', '.xls'):
            doc_type = 'xlsx'
        elif _ext in ('.txt',):
            doc_type = 'txt'
        elif _ext in ('.csv',):
            doc_type = 'csv'
        else:
            doc_type = None

        # Upsert SupplierDeclaration (Option B: material_id PK)
        existing = db.session.query(SupplierDeclaration).filter_by(material_id=material).first()
        if existing:
            existing.sku = sku
            existing.original_filename = fname
            existing.storage_filename = None
            existing.file_path = None
            existing.document_type = doc_type
            existing.supplier_name = supplier_name
            existing.description = description
            existing.upload_date = datetime.utcnow()
            existing.file_size = file_size
            existing.file_data = file_bytes
            decl = existing
        else:
            rec = SupplierDeclaration(
                material_id=material,
                sku=sku,
                original_filename=fname,
                storage_filename=None,
                file_path=None,
                document_type=doc_type,
                supplier_name=supplier_name,
                description=description,
                upload_date=datetime.utcnow(),
                file_size=file_size,
                file_data=file_bytes,
            )
            db.session.add(rec)
            decl = rec
        db.session.flush()  # Get ID before indexing
        
        # ✅ NEW: Trigger ChromaDB indexing after Postgres storage
        try:
            bom_row = db.session.query(PPWRBOM).filter_by(sku=sku, material_id=material).first()
            bom_metadata = {
                'material_name': bom_row.material_name if bom_row else '',
                'supplier_name': bom_row.supplier_name if bom_row else supplier_name or '',
                'component': bom_row.component if bom_row else '',
                'subcomponent': bom_row.subcomponent if bom_row else ''
            }
            
            # Call FastAPI indexing endpoint
            from fastapi_client import FASTAPI_BASE_URL
            from io import BytesIO
            import requests
            
            index_url = f"{FASTAPI_BASE_URL}/ppwr/index-declaration"
            index_files = {'file': (fname, BytesIO(file_bytes), 'application/pdf')}
            index_data = {
                'material_id': material,
                'sku': sku,
                'metadata': json.dumps(bom_metadata)
            }
            
            index_resp = requests.post(index_url, files=index_files, data=index_data, timeout=60)
            if index_resp.status_code == 200:
                index_result = index_resp.json()
                app.logger.info(f"✅ Indexed {index_result.get('chunks_created', 0)} chunks in ChromaDB for {material}")
            else:
                app.logger.warning(f"⚠️ ChromaDB indexing failed for {material}: HTTP {index_resp.status_code}")
        except Exception as e_index:
            app.logger.error(f"❌ ChromaDB indexing error for {material}: {e_index}")
            # Continue with upload even if indexing fails
        
        db.session.commit()
        flash('Declaration uploaded successfully', 'success')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"PPWR proxy: exception {e}", exc_info=True)
        flash(f"Error: {e}", 'danger')
    return redirect(url_for('ppwr_declarations_page'))


# Browser -> Frontend proxy: map declaration to material and persist in frontend DB,
# then forward mapping to backend FastAPI so both sides stay in sync.
@app.route('/api/ppwr/supplier-declarations/map', methods=['POST'])
def api_ppwr_map_supplier_declaration():
    try:
        # Accept form or json
        if request.is_json:
            payload = request.get_json(silent=True) or {}
            material_id = payload.get('material_id')
            apply_to_duplicates = payload.get('apply_to_duplicates', False)
            scope = payload.get('scope', 'sku')
        else:
            material_id = request.form.get('material_id')
            apply_to_duplicates = request.form.get('apply_to_duplicates', 'false')
            scope = request.form.get('scope', 'sku')

        if not material_id:
            return jsonify({'success': False, 'error': 'Missing material_id'}), 400

        mat = str(material_id).strip()
        # Always enforce one-to-one mapping; ignore any fan-out flags
        local_links_created = 0

        # Persist in frontend DB: validate BOM and set flag (mapping implicit under Option B)
        try:
            # Validate BOM contains this material_id
            try:
                if not PPWRBOM.query.filter_by(material_id=mat).first():
                    return jsonify({'success': False, 'error': f"PPWR BOM does not contain material_id '{mat}'"}), 400
            except Exception:
                return jsonify({'success': False, 'error': 'BOM validation failed'}), 500

            # Best-effort: mark ppwr_flag on BOM rows for this material
            try:
                rows = PPWRBOM.query.filter_by(material_id=mat).all()
                for r in rows:
                    r.ppwr_flag = True
                    db.session.add(r)
            except Exception:
                app.logger.debug('ppwr_flag set skipped')

            db.session.commit()
            local_links_created = 0

            # Best-effort: mark ppwr_flag on BOM rows for this material (optionally constrained by SKU)
            try:
                rows = PPWRBOM.query.filter_by(material_id=mat).all()
                for r in rows:
                    r.ppwr_flag = True
                    db.session.add(r)
            except Exception:
                app.logger.debug('ppwr_flag set skipped')

            db.session.commit()
        except Exception as e:
            db.session.rollback()
            app.logger.exception(f"Frontend mapping DB error: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500

        return jsonify({'success': True, 'links_created': local_links_created})
    except Exception as e:
        app.logger.exception(f"Mapping route failed: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/bom-uploads')
def bom_uploads_page():
    """Render a dedicated BOM Uploads page with drag/drop and client-side validation."""
    # Deprecated: redirect to dashboard; the standalone BOM upload page is removed.
    return redirect(url_for('index'))


# Admin: normalize duplicates by filename and archive older ones
@app.route('/api/admin/ppwr/cleanup-duplicate-filenames', methods=['POST'])
def api_admin_ppwr_cleanup_duplicate_filenames():
    try:
        # Group declarations by original_filename
        rows = SupplierDeclaration.query.filter_by(is_archived=False).order_by(SupplierDeclaration.id.desc()).all()
        groups = {}
        for r in rows:
            if not r.original_filename:
                continue
            groups.setdefault(r.original_filename, []).append(r)

        cleaned = 0
        for fname, decls in groups.items():
            # Pick newest by id
            decls_sorted = sorted(decls, key=lambda d: d.id or 0, reverse=True)
            newest = decls_sorted[0]
            target_mat = getattr(newest, 'material', None) or getattr(newest, 'material_id', None)

            # Unify materials to target_mat when available
            if target_mat:
                # Under Option B, mapping is implicit by material_id; no link updates needed.
                pass

            # Archive all but newest to reduce UI clutter
            for d in decls_sorted[1:]:
                if not getattr(d, 'is_archived', False):
                    d.is_archived = True
                    db.session.add(d)
            cleaned += 1

        db.session.commit()
        return jsonify({'success': True, 'groups_processed': cleaned})
    except Exception as e:
        db.session.rollback()
        app.logger.exception(f"Duplicate cleanup failed: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# Admin: hard purge all supplier declarations and links (fresh start)
@app.route('/api/admin/ppwr/purge-all', methods=['POST'])
def api_admin_ppwr_purge_all():
    try:
        # Remove links first, then declarations
        db.session.query(PPWRMaterialDeclarationLink).delete(synchronize_session=False)
        db.session.query(SupplierDeclaration).delete(synchronize_session=False)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        app.logger.exception(f"Purge failed: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== NEW PPWR BULK ACTION ROUTES ====================
# Import and register new PPWR bulk action routes (session-based checkboxes)
try:
    from ppwr_bulk_actions import register_ppwr_bulk_routes
    register_ppwr_bulk_routes(app, db, SupplierDeclaration, PPWRBOM, PPWRMaterialDeclarationLink, fastapi_assess_with_files)
    app.logger.info("✅ PPWR bulk action routes registered successfully")
except Exception as e:
    app.logger.warning(f"⚠️ Failed to register PPWR bulk action routes: {e}")


# Removed: separate Supplier Declarations page in favor of in-assessment uploads


@app.route('/api/bom/upload', methods=['POST'])
def api_bom_upload():
    """DEPRECATED: Legacy API endpoint for detailed BOM with component/subcomponent.
    
    Use the main /upload route instead, which populates route + ppwr_bom tables.
    """
    return jsonify({
        'success': False, 
        'error': 'This API endpoint is deprecated. Use the main /upload route (POST multipart with file) instead, which populates the simplified route + ppwr_bom structure.'
    }), 410
    try:
        file = request.files.get('file')
        if not file:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400

        filename = file.filename.lower()
        if filename.endswith('.csv'):
            df = pd.read_csv(file.stream, dtype=str)
        elif filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(file.stream, sheet_name=0, dtype=str)
        else:
            return jsonify({'success': False, 'error': 'Unsupported file type'}), 400

        # Normalize column names
        df.columns = [str(col).strip().replace('\n','').replace('\r','').replace(' ', '_').lower() for col in df.columns]

        # Accept a wider variety of header name variants (normalized to lower case, spaces -> underscores)
        required_cols = {
            'sku': [
                'sku', 'product_id', 'part_number',
                'sku_#', 'sku#', 'sku_number', 'sku_number'
            ],
            'product': [
                'product', 'product_name', 'product_desc', 'product_description'
            ],
            'component': [
                'component', 'comp', 'component_id',
                'component_#', 'component#', 'component_number', 'component_number'
            ],
            'component_description': [
                'component_description', 'comp_desc', 'component_desc', 'component description'
            ],
            'subcomponent': [
                'subcomponent', 'sub_comp', 'subcomponent_id',
                'sub_component_#', 'sub_component#', 'sub_component', 'sub component'
            ],
            'subcomponent_description': [
                'subcomponent_description', 'sub_comp_desc', 'subcomponent_desc', 'sub_component_description', 'sub component description'
            ],
            'material': [
                'material', 'mat_id', 'material_id',
                'material_#', 'material#', 'material_number', 'material_number'
            ],
            'material_name': [
                'material_name', 'mat_name', 'material_desc', 'material_description'
            ]
        }

        # Optionally accept chemical data columns in the BOM so we can persist them
        optional_cols = {
            'cas_number': ['cas', 'cas_number', 'cas_no', 'cas#'],
            'chemical_name': ['chemical_name', 'chem_name', 'chemical', 'substance', 'chemical name'],
            'concentration_ppm': ['concentration_ppm', 'concentration', 'conc_ppm', 'conc', 'ppm'],
            'supplier_name': ['supplier', 'supplier_name', 'supplier name', 'vendor', 'vendor_name'],
            'reference_doc': ['reference', 'reference_doc', 'reference_document', 'ref']
        }

        col_map = {}
        for target, possibles in required_cols.items():
            found = next((c for c in possibles if c in df.columns), None)
            if not found:
                return jsonify({'success': False, 'error': f'Missing required column: {target}'}), 400
            col_map[target] = found

        # If both a human-friendly 'material' column and an ID-like material column exist,
        # prefer the ID-like column for the internal material key so joins with `result` succeed.
        id_candidates = ['material_id', 'mat_id', 'material_code', 'material_no', 'material#', 'material_number', 'materialnumber', 'materialcode', 'matcode']
        for c in id_candidates:
            if c in df.columns:
                # override the material mapping to use ID-like column
                col_map['material'] = c
                break

        # Map optional columns if present
        optional_map = {}
        for target, possibles in optional_cols.items():
            found = next((c for c in possibles if c in df.columns), None)
            optional_map[target] = found
        app.logger.info(f"BOM API upload: column mapping: {col_map}")
        app.logger.info(f"BOM API upload: optional columns detected: {optional_map}")

        inserted = updated = skipped = 0
        audit_entries = []
        commit_time = datetime.utcnow()

        for index, row in df.iterrows():
            try:
                sku_val = str(row[col_map['sku']]).strip().split('.')[0]
                comp_val = str(row[col_map['component']]).strip().split('.')[0]
                subcomp_val = str(row[col_map['subcomponent']]).strip()

                # Normalize material value (handles multi-line cells like "Name\nID")
                material_raw = row[col_map['material']]
                material_val = _normalize_material_value(material_raw).split('.')[0]

                # If material is empty, try to derive from material_name column
                if not material_val and col_map.get('material_name'):
                    material_val = _normalize_material_value(row[col_map['material_name']]).split('.')[0]

                # Support multiple materials in one cell (comma/semicolon/pipe/slash separated)
                material_list = [m.strip() for m in re.split(r'[;,|/]+', material_val) if m.strip()] if material_val else []
                if not material_list:
                    # nothing we can do for this row
                    skipped += 1
                    audit_entries.append({'sku': sku_val, 'material': material_val, 'action': 'skip', 'details': 'missing primary key'})
                    continue

                # For each material in the (possibly split) list, insert/update BOM row
                for material_val_single in material_list:
                    material_val_single = material_val_single.split('.')[0]
                    try:
                        existing = db.session.query(PFASBOM).filter_by(
                            sku=sku_val, component=comp_val, subcomponent=subcomp_val, material=material_val_single
                        ).first()

                        if existing:
                            existing.product = str(row[col_map['product']]).strip()
                            # Update non-key fields
                            existing.component_description = str(row[col_map['component_description']]).strip() if col_map.get('component_description') and col_map.get('component_description') in row else existing.component_description
                            existing.subcomponent_description = str(row[col_map['subcomponent_description']]).strip() if col_map.get('subcomponent_description') and col_map.get('subcomponent_description') in row else existing.subcomponent_description
                            existing.material_name = str(row[col_map['material_name']]).strip() if col_map.get('material_name') and col_map.get('material_name') in row else existing.material_name
                            existing.uploaded_at = commit_time
                            updated += 1
                            action = 'update'
                        else:
                            bom_item = PFASBOM(
                                sku=sku_val,
                                material=material_val_single,
                                product=str(row[col_map['product']]).strip() if col_map.get('product') else None,
                                component=comp_val,
                                component_description=str(row[col_map['component_description']]).strip() if col_map.get('component_description') else None,
                                subcomponent=subcomp_val,
                                subcomponent_description=str(row[col_map['subcomponent_description']]).strip() if col_map.get('subcomponent_description') else None,
                                material_name=str(row[col_map['material_name']]).strip() if col_map.get('material_name') else None,
                                portal_name='SAP',
                                region='Global',
                                assessment=','.join(request.form.getlist('assessments')) or 'PFAS',
                                uploaded_at=commit_time
                            )
                            db.session.add(bom_item)
                            inserted += 1
                            action = 'insert'

                        # Append an audit entry per material
                        audit_entries.append({
                            'sku': sku_val,
                            'product': str(row[col_map['product']]).strip() if col_map.get('product') else None,
                            'component': comp_val,
                            'component_description': str(row[col_map['component_description']]).strip() if col_map.get('component_description') else None,
                            'subcomponent': subcomp_val,
                            'subcomponent_description': str(row[col_map['subcomponent_description']]).strip() if col_map.get('subcomponent_description') else None,
                            'material': material_val_single,
                            'material_name': str(row[col_map['material_name']]).strip() if col_map.get('material_name') else None,
                            'action': action,
                            'uploaded_at': commit_time.isoformat()
                        })
                    except Exception as e:
                        # per-material error
                        db.session.rollback()
                        skipped += 1
                        audit_entries.append({'sku': sku_val, 'material': material_val_single, 'action': 'error', 'details': str(e)})
                        continue

                # ---- Optional: persist chemical data if present in the BOM row ----
                try:
                    def _get(col_key):
                        name = optional_map.get(col_key)
                        if not name:
                            return None
                        val = row[name]
                        if pd.isna(val):
                            return None
                        s = str(val).strip()
                        return s if s != '' else None

                    # Read chemical fields once per-row
                    cas_val = _get('cas_number')
                    chem_val = _get('chemical_name')
                    supplier_val = _get('supplier_name')
                    ref_val = _get('reference_doc')

                    conc_val = None
                    conc_raw = _get('concentration_ppm')
                    if conc_raw is not None:
                        try:
                            # store as Decimal to match Numeric column
                            conc_val = Decimal(conc_raw)
                        except (InvalidOperation, ValueError):
                            try:
                                conc_val = Decimal(str(float(conc_raw)))
                            except Exception:
                                conc_val = None

                    # Only upsert if we have at least one chemical-related field
                    if any([cas_val, chem_val, conc_val is not None, supplier_val, ref_val]):
                        # Upsert per-material entry (material_val_single) — ensure the
                        # chemical row matches the exact material ID inserted above.
                        for material_val_single in material_list:
                            m_id = material_val_single
                            existing_mat = db.session.query(PFASMaterialChemicals).filter_by(material_id=m_id).first()
                            if existing_mat:
                                if cas_val:
                                    existing_mat.cas_number = cas_val
                                if chem_val:
                                    existing_mat.chemical_name = chem_val
                                if conc_val is not None:
                                    existing_mat.concentration_ppm = conc_val
                                if supplier_val:
                                    existing_mat.supplier_name = supplier_val
                                if ref_val:
                                    existing_mat.reference_doc = ref_val
                            else:
                                new_mat = PFASMaterialChemicals(
                                    material_id=m_id,
                                    cas_number=cas_val,
                                    material_name=str(row[col_map['material_name']]).strip() if col_map.get('material_name') else None,
                                    chemical_name=chem_val,
                                    concentration_ppm=conc_val,
                                    supplier_name=supplier_val,
                                    reference_doc=ref_val
                                )
                                db.session.add(new_mat)

                            # mark BOM row(s) as having chemical data
                            try:
                                bom_rows = db.session.query(PFASBOM).filter_by(sku=sku_val, component=comp_val, subcomponent=subcomp_val, material=m_id).all()
                                for bom_row in bom_rows:
                                    bom_row.flag = True
                            except Exception:
                                pass

                        # Log what we wrote/updated for this material row for debugging
                        try:
                            app.logger.info(f"BOM API upload: upserted chemical for materials={material_list} cas={cas_val} chem={chem_val} conc={conc_val} supplier={supplier_val} ref={ref_val}")
                        except Exception:
                            pass
                except Exception:
                    # Non-fatal: don't break entire upload if chemical upsert fails
                    app.logger.debug(f"Optional chemical upsert failed for row {index}", exc_info=True)

            except Exception as e:
                db.session.rollback()
                skipped += 1
                audit_entries.append({'sku': None, 'material': None, 'action': 'error', 'details': str(e)})

        db.session.commit()

        # Persist audit rows
        try:
            for a in audit_entries:
                audit = PFASBOMAudit(
                    sku=a.get('sku'),
                    product=a.get('product'),
                    component=a.get('component'),
                    component_description=a.get('component_description'),
                    subcomponent=a.get('subcomponent'),
                    subcomponent_description=a.get('subcomponent_description'),
                    material=a.get('material'),
                    material_name=a.get('material_name'),
                    action=a.get('action'),
                    details=a.get('details', ''),
                    uploaded_at=commit_time
                )
                db.session.add(audit)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            app.logger.warning(f"Failed to write audit rows: {e}")

        return jsonify({'success': True, 'inserted': inserted, 'updated': updated, 'skipped': skipped, 'uploaded_at': commit_time.isoformat()}), 200

    except Exception as e:
        app.logger.error(f"BOM upload API failure: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/upload', methods=['POST'])
def upload_bom():
    app.logger.info("Starting BOM upload (route + ppwr_bom)")
    # Clear any failed transaction state
    try:
        db.session.rollback()
    except Exception:
        pass

    file = request.files.get('file')
    if not file:
        flash("No file uploaded.", "danger")
        return redirect(url_for('index'))

    # Determine route from assessment selection
    route_raw = (request.form.get('assessment') or ",".join(request.form.getlist('assessments')) or 'PPWR').strip().lower()
    route_val = 'ppwr' if 'ppwr' in route_raw else 'pfas'

    try:
        filename = (file.filename or '').lower()
        _, _ext = os.path.splitext(filename)
        # Reject supplier declarations (PDFs) in BOM upload
        if _ext in ALLOWED_DECL_EXT and _ext == '.pdf':
            app.logger.info(f"PDF detected in BOM upload; advising user to use Assessment Upload instead: {filename}")
            flash("Unsupported file for BOM: Please upload CSV/Excel for BOM. Use the PPWR Assessment 'Upload' button per row to add supplier declarations.", "warning")
            return redirect(url_for('index'))

        # Read raw bytes
        file_bytes = file.read()
        buf = BytesIO(file_bytes)

        # Parse BOM to extract SKU and materials
        if filename.endswith('.csv'):
            df = pd.read_csv(buf, dtype=str)
        elif filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(buf, sheet_name=0, dtype=str)
        else:
            flash("Unsupported BOM file type. Please upload CSV or Excel.", "danger")
            return redirect(url_for('index'))

        # Normalize headers
        df.columns = [str(c).strip().replace('\n','').replace('\r','').replace(' ','_').lower() for c in df.columns]
        
        # Find required columns
        sku_col = next((c for c in ['sku','product_id','part_number'] if c in df.columns), None)
        material_col = next((c for c in ['material', 'mat_id', 'material_id', 'material_#', 'material#', 'material_number'] if c in df.columns), None)
        
        if not sku_col or not material_col or df.empty:
            flash("Missing required SKU/Material columns or empty file.", "danger")
            return redirect(url_for('index'))

        sku_val = str(df.iloc[0][sku_col]).strip()
        now = datetime.utcnow()

        # 1. Upsert route table with (sku, route)
        route_rec = db.session.query(Route).filter_by(sku=sku_val).first()
        if route_rec:
            route_rec.route = route_val
        else:
            route_rec = Route(sku=sku_val, route=route_val)
            db.session.add(route_rec)

        # 2. Parse materials from BOM and populate ppwr_bom
        material_name_col = next((c for c in ['material_name', 'mat_name', 'material_desc'] if c in df.columns), None)
        supplier_col = next((c for c in ['supplier', 'supplier_name', 'vendor'] if c in df.columns), None)
        
        materials_inserted = 0
        materials_updated = 0
        
        for _, row in df.iterrows():
            try:
                material_raw = row[material_col]
                material_val = _normalize_material_value(material_raw).split('.')[0]
                
                if not material_val:
                    continue
                
                # Support multiple materials in one cell (comma/semicolon/pipe/slash separated)
                material_list = [m.strip() for m in re.split(r'[;,|/]+', material_val) if m.strip()]
                
                for mat_id in material_list:
                    mat_id = mat_id.split('.')[0]
                    
                    # Get optional fields
                    mat_name = str(row[material_name_col]).strip() if material_name_col and material_name_col in row else None
                    supplier_name = str(row[supplier_col]).strip() if supplier_col and supplier_col in row else None
                    
                    # Upsert ppwr_bom row
                    existing = db.session.query(PPWRBOM).filter_by(material_id=mat_id).first()
                    if existing:
                        existing.sku = sku_val
                        existing.material_name = mat_name or existing.material_name
                        existing.supplier_name = supplier_name or existing.supplier_name
                        existing.uploaded_at = now
                        materials_updated += 1
                    else:
                        new_mat = PPWRBOM(
                            material_id=mat_id,
                            sku=sku_val,
                            material_name=mat_name,
                            supplier_name=supplier_name,
                            ppwr_flag=False,
                            uploaded_at=now
                        )
                        db.session.add(new_mat)
                        materials_inserted += 1
                        
            except Exception as row_err:
                app.logger.warning(f"Failed to parse material from row: {row_err}")
                continue

        db.session.commit()
        app.logger.info(f"BOM upload: route upserted, {materials_inserted} materials inserted, {materials_updated} updated")

        # Dashboard highlight and success message
        session['last_upload'] = {'timestamp': now.strftime('%Y-%m-%d %H:%M:%S UTC'), 'skus': [sku_val]}
        flash(f"BOM_UPLOADED_SUCCESSFULLY: SKU={sku_val} route={route_val}, {materials_inserted + materials_updated} materials processed", "success")
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Upload failed: {e}", exc_info=True)
        flash("❌ Error processing file.", "danger")

    return redirect(url_for('index'))


@app.route('/debug-material/<material_id>')
def debug_material(material_id):
    """Debug endpoint: return the stored chemical row for a material_id (from `result` table).

    Use this to confirm whether the BOM upload wrote a `result` row for the given material id.
    """
    try:
        app.logger.info(f"Debug material lookup for: {material_id}")
        mat = db.session.query(PFASMaterialChemicals).filter_by(material_id=material_id).first()
        if not mat:
            return jsonify({'found': False, 'material_id': material_id}), 200

        conc = None
        try:
            conc = float(mat.concentration_ppm) if mat.concentration_ppm is not None else None
        except Exception:
            conc = str(mat.concentration_ppm) if mat.concentration_ppm is not None else None

        return jsonify({
            'found': True,
            'material_id': mat.material_id,
            'material_name': mat.material_name,
            'cas_number': mat.cas_number,
            'chemical_name': mat.chemical_name,
            'concentration_ppm': conc,
            'supplier_name': mat.supplier_name,
            'reference_doc': mat.reference_doc
        }), 200
    except Exception as e:
        app.logger.error(f"Debug material error: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/debug-bom/<sku>')
def debug_bom(sku):
    """DEPRECATED: Debug route for old PFASBOM structure. Use ppwr_bom now."""
    return jsonify({'error': 'PFASBOM table replaced by route + ppwr_bom. This debug route is deprecated.'}), 410


@app.route('/upload-supplier-declaration', methods=['POST'])
def upload_supplier_declaration_frontend():
    """Accept a supplier declaration file and store it in Postgres (same DB as BOM).

    Expects multipart/form-data with:
      - file: the uploaded file
      - sku: (optional) product SKU to attach
      - material: (optional) material id to attach
      - metadata: (optional) JSON string with extra metadata
    Returns JSON with success and created metadata.
    """
    try:
        f = request.files.get('file')
        if not f:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400

        fname = f.filename
        if not _allowed_decl(fname):
            return jsonify({'success': False, 'error': 'Unsupported file extension'}), 400

        sku = request.form.get('sku')
        material = request.form.get('material')
        supplier_name = request.form.get('supplier_name')
        description = request.form.get('description')

        # If material is not provided, infer from filename prefix (ID-like token)
        if not material:
            try:
                base = os.path.splitext(fname)[0]
                tokens = [t for t in re.split(r'[\s_\-]+', base) if t]
                if tokens:
                    cand = tokens[0]
                    if re.match(r'^[A-Za-z0-9_\-]{1,40}$', cand):
                        material = cand
            except Exception:
                pass

        # Parse optional metadata JSON
        metadata = None
        metadata_raw = request.form.get('metadata')
        if metadata_raw:
            try:
                metadata = json.loads(metadata_raw)
            except Exception:
                metadata = None

        # Read file bytes for DB storage
        file_bytes = f.read()
        file_size = len(file_bytes) if file_bytes is not None else None
        # Derive simple document type
        _ext = os.path.splitext(fname.lower())[1]
        if _ext in ('.pdf',):
            doc_type = 'pdf'
        elif _ext in ('.docx', '.doc'):
            doc_type = 'docx'
        elif _ext in ('.xlsx', '.xls'):
            doc_type = 'xlsx'
        elif _ext in ('.txt',):
            doc_type = 'txt'
        elif _ext in ('.csv',):
            doc_type = 'csv'
        else:
            doc_type = None

        # Create/Update ORM entity for current SupplierDeclaration model (Option B: material_id PK)
        # Validate BOM existence; warn but allow upload to proceed so UI can demonstrate storage
        try:
            if material:
                if sku:
                    exists = db.session.query(PFASBOM).filter_by(sku=sku, material=material).first()
                else:
                    exists = db.session.query(PFASBOM).filter_by(material=material).first()
                if not exists:
                    app.logger.warning(f"Supplier declaration upload: no BOM match for sku={sku} material={material}; proceeding with upload")
        except Exception as _e:
            app.logger.exception("Supplier declaration upload: BOM validation error; proceeding with upload")

        existing = SupplierDeclaration.query.filter_by(material_id=material).first() if material else None
        if existing:
            existing.sku = sku
            existing.original_filename = fname
            existing.storage_filename = None
            existing.file_path = None
            existing.document_type = doc_type
            existing.upload_date = datetime.utcnow()
            existing.metadata_json = metadata
            existing.file_size = file_size
            existing.file_data = file_bytes
            existing.supplier_name = supplier_name
            existing.description = description
            decl = existing
        else:
            decl = SupplierDeclaration(
                material_id=material,
                sku=sku,
                original_filename=fname,
                storage_filename=None,
                file_path=None,
                document_type=doc_type,
                upload_date=datetime.utcnow(),
                metadata_json=metadata,
                file_size=file_size,
                file_data=file_bytes,
                supplier_name=supplier_name,
                description=description,
            )
            db.session.add(decl)
        try:
            db.session.commit()
        except Exception as commit_err:
            db.session.rollback()
            # Fallback: some legacy schemas enforce NOT NULL on columns like
            # document_type/original_filename/storage_filename/file_path.
            # Insert via raw SQL including those fields with benign defaults.
            try:
                # Fallback insert supporting legacy columns (now uses material_id PK)
                sql = text(
                    """
                    INSERT INTO supplier_declarations
                        (material_id, sku, original_filename, storage_filename, file_path, document_type, upload_date,
                         metadata_json, file_size, file_data, supplier_name, description)
                    VALUES
                        (:material_id, :sku, :original_filename, :storage_filename, :file_path, :document_type, :upload_date,
                         CAST(:metadata_json AS JSON), :file_size, :file_data, :supplier_name, :description)
                    ON CONFLICT (material_id) DO UPDATE SET
                        sku = EXCLUDED.sku,
                        original_filename = EXCLUDED.original_filename,
                        storage_filename = EXCLUDED.storage_filename,
                        file_path = EXCLUDED.file_path,
                        document_type = EXCLUDED.document_type,
                        upload_date = EXCLUDED.upload_date,
                        metadata_json = EXCLUDED.metadata_json,
                        file_size = EXCLUDED.file_size,
                        file_data = EXCLUDED.file_data,
                        supplier_name = EXCLUDED.supplier_name,
                        description = EXCLUDED.description
                    RETURNING material_id
                    """
                )
                params = {
                    'material_id': material,
                    'sku': sku,
                    'original_filename': fname,
                    'storage_filename': None,
                    'file_path': None,
                    'document_type': doc_type,
                    'upload_date': datetime.utcnow(),
                    'metadata_json': json.dumps(metadata) if isinstance(metadata, (dict, list)) else None,
                    'file_size': file_size,
                    'file_data': file_bytes,
                    'supplier_name': supplier_name,
                    'description': description,
                }
                res = db.session.execute(sql, params)
                row = res.fetchone()
                db.session.commit()
                new_id = row[0] if row else None
                return jsonify({
                    'success': True,
                    'material_id': new_id,
                    'filename': fname,
                    'sku': sku,
                    'material': material,
                    'file_size': file_size
                }), 201
            except Exception as e2:
                app.logger.error(f"Supplier declaration upload fallback insert failed: {e2}", exc_info=True)
                raise commit_err

        return jsonify({
            'success': True,
            'material_id': decl.material_id,
            'filename': decl.original_filename,
            'sku': decl.sku,
            'material': decl.material_id,
            'supplier_name': decl.supplier_name,
            'description': decl.description,
            'file_size': decl.file_size
        }), 201

    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        app.logger.error(f"Supplier declaration upload failed: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/manual-upsert-chemical', methods=['POST'])
def manual_upsert_chemical():
    """Admin helper: manually insert or update a PFASMaterialChemicals row.

    Accepts JSON body with keys:
      - material_id (required)
      - cas_number
      - chemical_name
      - concentration_ppm (number)
      - supplier_name
      - reference_doc

    This is intended as a small admin/debug helper to populate the `result` table
    when automated ingestion is not available. It performs an upsert and returns
    the created/updated row.
    """
    try:
        payload = request.get_json(force=True)
        if not payload:
            return jsonify({'success': False, 'error': 'Missing JSON body'}), 400

        material_id = payload.get('material_id')
        if not material_id:
            return jsonify({'success': False, 'error': 'material_id is required'}), 400

        cas = payload.get('cas_number')
        chem = payload.get('chemical_name')
        conc = payload.get('concentration_ppm')
        supplier = payload.get('supplier_name')
        ref = payload.get('reference_doc')

        # Normalize concentration if provided
        conc_val = None
        if conc is not None:
            try:
                from decimal import Decimal
                conc_val = Decimal(str(conc))
            except Exception:
                conc_val = None

        # Upsert into PFASMaterialChemicals
        existing = db.session.query(PFASMaterialChemicals).filter_by(material_id=material_id).first()
        if existing:
            if cas:
                existing.cas_number = cas
            if chem:
                existing.chemical_name = chem
            if conc_val is not None:
                existing.concentration_ppm = conc_val
            if supplier:
                existing.supplier_name = supplier
            if ref:
                existing.reference_doc = ref
            action = 'updated'
        else:
            new_row = PFASMaterialChemicals(
                material_id=material_id,
                cas_number=cas,
                material_name=payload.get('material_name'),
                chemical_name=chem,
                concentration_ppm=conc_val,
                supplier_name=supplier,
                reference_doc=ref
            )
            db.session.add(new_row)
            action = 'inserted'

        db.session.commit()

        return jsonify({'success': True, 'action': action, 'material_id': material_id}), 200
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"manual_upsert_chemical failed: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/assessment-upload', methods=['POST'])
def api_assessment_upload():
    """Compatibility endpoint for the assessment UI upload buttons.

    Re-uses the existing supplier-declaration upload logic so the UI can POST
    files to /api/assessment-upload and receive the same JSON response.
    """
    app.logger.info("/api/assessment-upload called")
    # Forward to existing handler which expects multipart/form-data
    return upload_supplier_declaration_frontend()


@app.route('/supplier-declaration/<int:decl_id>/download')
def download_supplier_declaration(decl_id):
    try:
        return jsonify({'error': 'Download by numeric id is deprecated. Use /supplier-declaration/<material_id>/download'}), 410
    except Exception as e:
        app.logger.error(f"Download failed for declaration {decl_id}: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/supplier-declaration/<material_id>/download')
def download_supplier_declaration_by_material(material_id):
    try:
        decl = db.session.query(SupplierDeclaration).filter_by(material_id=str(material_id)).first()
        if not decl:
            return jsonify({'error': 'Not found'}), 404

        if getattr(decl, 'file_data', None):
            return send_file(
                BytesIO(decl.file_data),
                as_attachment=True,
                download_name=getattr(decl, 'original_filename', 'document.bin'),
                mimetype='application/pdf' if getattr(decl, 'document_type', None) == 'pdf' else 'application/octet-stream'
            )

        file_path = getattr(decl, 'file_path', None)
        if file_path and os.path.exists(file_path):
            return send_file(file_path, as_attachment=True, download_name=getattr(decl, 'original_filename', 'document.bin'))

        return jsonify({'error': 'File missing'}), 404
    except Exception as e:
        app.logger.error(f"Download failed for material {material_id}: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/supplier-declarations/<sku>')
def list_supplier_declarations(sku):
    try:
        include_archived = request.args.get('include_archived', '0') in ('1', 'true', 'True')
        # Use fields defined on the current `SupplierDeclaration` model in models.py
        q = db.session.query(SupplierDeclaration).filter_by(sku=sku)
        if not include_archived:
            # Filter out archived if the column exists
            try:
                q = q.filter(SupplierDeclaration.is_archived == False)  # noqa: E712
            except Exception:
                pass
        rows = q.order_by(SupplierDeclaration.upload_date.desc()).all()

        items = []
        for r in rows:
            # Provide both legacy keys and UI-expected keys for compatibility
            uploaded_iso = r.upload_date.isoformat() if getattr(r, 'upload_date', None) else None
            items.append({
                # Core (no numeric id under Option B)
                'sku': r.sku,
                'material_id': getattr(r, 'material_id', None),
                'material': getattr(r, 'material_id', None),
                'file_size': getattr(r, 'file_size', None),
                'metadata': getattr(r, 'metadata_json', None),
                'is_archived': getattr(r, 'is_archived', False),
                # Current UI expects these keys
                'original_filename': getattr(r, 'original_filename', None),
                'document_type': getattr(r, 'document_type', None),
                'upload_date': uploaded_iso,
                'supplier_name': getattr(r, 'supplier_name', None),
                'description': getattr(r, 'description', None),
                # Back-compat keys still used elsewhere
                'filename': getattr(r, 'original_filename', None),
                'stored_path': getattr(r, 'file_path', None),
                'content_type': getattr(r, 'document_type', None),
                'uploaded_at': uploaded_iso,
            })

        # Return under 'declarations' key for the frontend JS to consume
        return jsonify({'success': True, 'declarations': items}), 200
    except Exception as e:
        app.logger.error(f"List supplier declarations failed: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# Strict mapping status: join declarations to BOM by sku+material
@app.route('/api/document-mapping/<sku>', methods=['GET'])
def api_document_mapping(sku):
    try:
        # Fetch declarations for the SKU
        decls = db.session.query(SupplierDeclaration).filter_by(sku=sku).all()
        # Build BOM material set for the SKU
        bom_rows = db.session.query(PFASBOM.material).filter_by(sku=sku).all()
        bom_set = set([r.material for r in bom_rows])

        items = []
        for d in decls:
            mat = getattr(d, 'material_id', None)
            is_mapped = bool(mat) and mat in bom_set
            items.append({
                'id': None,
                'filename': getattr(d, 'original_filename', None),
                'document_type': getattr(d, 'document_type', None),
                'material': mat,
                'is_mapped': is_mapped,
                'mapped_to': mat if is_mapped else None,
                'upload_date': d.upload_date.isoformat() if getattr(d, 'upload_date', None) else None
            })
        return jsonify({'success': True, 'sku': sku, 'mappings': items}), 200
    except Exception as e:
        app.logger.error(f"document-mapping failed: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


# New: UI-compatible upload endpoint supporting multiple files
@app.route('/api/supplier-declarations/upload', methods=['POST'])
def api_supplier_declarations_upload():
    """Accept multiple supplier declaration files and store them in Postgres.

    Form fields:
      - files: one or more files
      - sku: required target SKU
      - material: optional material id
      - supplier_name: optional supplier name
      - description: optional description
    """
    try:
        sku = request.form.get('sku')
        if not sku:
            return jsonify({'error': 'Missing sku'}), 400

        material = request.form.get('material')
        supplier_name = request.form.get('supplier_name')
        description = request.form.get('description')

        files = request.files.getlist('files')
        if not files:
            return jsonify({'error': 'No files uploaded'}), 400

        uploaded = []
        commit_time = datetime.utcnow()
        errors = []
        for f in files:
            try:
                fname = f.filename
                if not _allowed_decl(fname):
                    errors.append({'filename': fname, 'error': 'Unsupported file extension'})
                    continue

                # Infer material from filename if not provided
                mat_val = material or ''
                if not mat_val:
                    try:
                        base = os.path.splitext(fname)[0]
                        tokens = [t for t in re.split(r'[\s_\-]+', base) if t]
                        if tokens:
                            cand = tokens[0]
                            if re.match(r'^[A-Za-z0-9_\-]{1,40}$', cand):
                                mat_val = cand
                    except Exception:
                        mat_val = ''

                # Validate BOM existence for SKU + material
                if not mat_val:
                    errors.append({'filename': fname, 'error': 'Material could not be inferred from filename'})
                    continue
                
                # ✅ BACKEND VALIDATION: Check BOM row exists
                try:
                    bom_row = db.session.query(PPWRBOM).filter_by(sku=sku, material_id=mat_val).first()
                    if not bom_row:
                        errors.append({
                            'filename': fname, 
                            'error': f"No BOM row for SKU '{sku}' and material '{mat_val}'"
                        })
                        continue
                    
                    # ✅ BACKEND VALIDATION: Check filename format MaterialID_MaterialName
                    expected_mat_name = (bom_row.material_name or '').strip()
                    if expected_mat_name:
                        fname_lower = fname.lower()
                        expected_lower = expected_mat_name.lower()
                        base_name = os.path.splitext(fname)[0].lower()
                        mat_val_lower = mat_val.lower()
                        
                        # Check if filename follows MaterialID_MaterialName pattern
                        if not base_name.startswith(mat_val_lower + '_'):
                            errors.append({
                                'filename': fname,
                                'error': f"Filename must start with '{mat_val}_'. Expected format: {mat_val}_{expected_mat_name}.pdf"
                            })
                            continue
                        
                        # Check if material name is present after material ID
                        if expected_lower not in fname_lower:
                            errors.append({
                                'filename': fname,
                                'error': f"Filename must contain material name. Expected format: {mat_val}_{expected_mat_name}.pdf"
                            })
                            continue
                except Exception as _e:
                    errors.append({'filename': fname, 'error': f'Filename validation failed: {str(_e)}'})
                    continue

                file_bytes = f.read()
                file_size = len(file_bytes) if file_bytes is not None else None
                _ext = os.path.splitext(fname.lower())[1]
                if _ext in ('.pdf',):
                    doc_type = 'pdf'
                elif _ext in ('.docx', '.doc'):
                    doc_type = 'docx'
                elif _ext in ('.xlsx', '.xls'):
                    doc_type = 'xlsx'
                elif _ext in ('.txt',):
                    doc_type = 'txt'
                elif _ext in ('.csv',):
                    doc_type = 'csv'
                else:
                    doc_type = None

                existing = SupplierDeclaration.query.filter_by(material_id=mat_val).first()
                if existing:
                    existing.sku = sku
                    existing.original_filename = fname
                    existing.storage_filename = None
                    existing.file_path = None
                    existing.document_type = doc_type
                    existing.supplier_name = supplier_name
                    existing.description = description
                    existing.upload_date = commit_time
                    existing.file_size = file_size
                    existing.file_data = file_bytes
                    decl = existing
                else:
                    decl = SupplierDeclaration(
                        material_id=mat_val,
                        sku=sku,
                        original_filename=fname,
                        storage_filename=None,
                        file_path=None,
                        document_type=doc_type,
                        supplier_name=supplier_name,
                        description=description,
                        upload_date=commit_time,
                        file_size=file_size,
                        file_data=file_bytes,
                    )
                    db.session.add(decl)
                db.session.flush()
                
                # ✅ NEW: Trigger ChromaDB indexing after Postgres storage
                try:
                    bom_metadata = {
                        'material_name': bom_row.material_name if bom_row else '',
                        'supplier_name': bom_row.supplier_name if bom_row else supplier_name or '',
                        'component': bom_row.component if bom_row else '',
                        'subcomponent': bom_row.subcomponent if bom_row else ''
                    }
                    
                    from fastapi_client import FASTAPI_BASE_URL
                    from io import BytesIO
                    import requests
                    
                    index_url = f"{FASTAPI_BASE_URL}/ppwr/index-declaration"
                    index_files = {'file': (fname, BytesIO(file_bytes), 'application/pdf')}
                    index_data = {
                        'material_id': mat_val,
                        'sku': sku,
                        'metadata': json.dumps(bom_metadata)
                    }
                    
                    index_resp = requests.post(index_url, files=index_files, data=index_data, timeout=60)
                    if index_resp.status_code == 200:
                        index_result = index_resp.json()
                        app.logger.info(f"✅ Indexed {index_result.get('chunks_created', 0)} chunks in ChromaDB for {mat_val}")
                    else:
                        app.logger.warning(f"⚠️ ChromaDB indexing failed for {mat_val}: HTTP {index_resp.status_code}")
                except Exception as e_index:
                    app.logger.error(f"❌ ChromaDB indexing error for {mat_val}: {e_index}")
                    # Continue with upload even if indexing fails
                
                uploaded.append({
                    'material_id': decl.material_id,
                    'original_filename': decl.original_filename,
                    'upload_date': decl.upload_date.isoformat() if decl.upload_date else None
                })
            except Exception as e:
                db.session.rollback()
                errors.append({'filename': f.filename if f else None, 'error': str(e)})
                # continue with next file
        # Try to commit all successful inserts
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e), 'uploaded': uploaded, 'errors': errors, 'total_uploaded': 0, 'total_errors': len(files)}), 500

        status = 201 if uploaded else 400
        return jsonify({
            'success': True if uploaded else False,
            'uploaded': uploaded,
            'errors': errors,
            'total_uploaded': len(uploaded),
            'total_errors': len(errors),
            'upload_time': commit_time.isoformat() + 'Z'
        }), status
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'error': str(e)}), 500


# New: UI-compatible download endpoint
@app.route('/api/supplier-declarations/download/<material_id>')
def api_supplier_declarations_download(material_id):
    return download_supplier_declaration_by_material(material_id)

# New: Proxy upload to backend PPWR storage (single file)
@app.route('/api/ppwr/supplier-declarations/upload', methods=['POST'])
def api_ppwr_supplier_declarations_upload():
    """Deprecated: Storage moved to Flask-only. Use /api/supplier-declarations/upload."""
    return jsonify({'success': False, 'error': 'Supplier declaration storage moved to Flask. Use /api/supplier-declarations/upload'}), 410


# (Removed duplicate mapping route)


# New: UI-compatible delete endpoint
def _archive_declaration_by_id(decl_id: int, req_sku: str | None):
    """Deprecated: numeric id no longer used under Option B."""
    return False, 410, 'Archive by id is deprecated; archive by material_id instead'
    # Try ORM update
    try:
        setattr(decl, 'is_archived', True)
        db.session.commit()
        return True, 200, 'Document archived'
    except Exception as e:
        app.logger.warning(f"ORM archive failed for decl_id={decl_id}: {e}. Falling back to raw SQL.")
        try:
            db.session.rollback()
        except Exception:
            pass
        # Raw SQL fallback
        try:
            if req_sku:
                result = db.session.execute(
                    text('UPDATE supplier_declarations SET is_archived = TRUE WHERE id = :id AND sku = :sku'),
                    {'id': decl_id, 'sku': req_sku}
                )
            else:
                result = db.session.execute(
                    text('UPDATE supplier_declarations SET is_archived = TRUE WHERE id = :id'),
                    {'id': decl_id}
                )
            db.session.commit()
            if hasattr(result, 'rowcount') and result.rowcount == 0:
                return False, 404, 'Not found or SKU mismatch'
            return True, 200, 'Document archived'
        except Exception as e2:
            try:
                db.session.rollback()
            except Exception:
                pass
            app.logger.exception(f"Raw SQL archive also failed for decl_id={decl_id}")
            return False, 500, str(e2)


@app.route('/api/supplier-declarations/<int:decl_id>', methods=['DELETE'])
def api_supplier_declarations_delete(decl_id):
    try:
        # Ensure we aren't in a failed transaction state from a prior error
        try:
            db.session.rollback()
        except Exception:
            pass

        # Optional safety: require sku match if provided by client
        req_sku = request.args.get('sku')

        # Soft-delete (archive) by default to avoid FK violations and keep audit trails
        ok, code, msg = _archive_declaration_by_id(decl_id, req_sku)
        if not ok:
            return jsonify({'error': msg}), code
        return jsonify({'success': True, 'message': msg}), 200
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        app.logger.exception(f"Delete supplier declaration failed for id={decl_id}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/supplier-declarations/<int:decl_id>/restore', methods=['POST'])
def api_supplier_declarations_restore(decl_id):
    """Unarchive a supplier declaration (set is_archived=false). Accepts optional sku for safety."""
    try:
        try:
            db.session.rollback()
        except Exception:
            pass

        req_sku = request.args.get('sku')
        decl = db.session.query(SupplierDeclaration).filter_by(id=decl_id).first()
        if not decl:
            return jsonify({'error': 'Not found'}), 404
        if req_sku and decl.sku != req_sku:
            return jsonify({'error': 'SKU mismatch for restore request'}), 400

        try:
            setattr(decl, 'is_archived', False)
            db.session.commit()
            return jsonify({'success': True, 'message': 'Document restored'}), 200
        except Exception as e:
            app.logger.warning(f"ORM restore failed for decl_id={decl_id}: {e}. Falling back to raw SQL.")
            try:
                db.session.rollback()
            except Exception:
                pass
            try:
                if req_sku:
                    result = db.session.execute(
                        text('UPDATE supplier_declarations SET is_archived = FALSE WHERE id = :id AND sku = :sku'),
                        {'id': decl_id, 'sku': req_sku}
                    )
                else:
                    result = db.session.execute(
                        text('UPDATE supplier_declarations SET is_archived = FALSE WHERE id = :id'),
                        {'id': decl_id}
                    )
                db.session.commit()
                if hasattr(result, 'rowcount') and result.rowcount == 0:
                    return jsonify({'error': 'Not found or SKU mismatch'}), 404
                return jsonify({'success': True, 'message': 'Document restored'}), 200
            except Exception as e2:
                try:
                    db.session.rollback()
                except Exception:
                    pass
                app.logger.exception(f"Raw SQL restore also failed for decl_id={decl_id}")
                return jsonify({'error': str(e2)}), 500
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        app.logger.exception(f"Restore supplier declaration failed for id={decl_id}")
        return jsonify({'error': str(e)}), 500


def pfas_assessment(sku):
    try:
        # Ensure any prior failed transaction is cleared
        try:
            db.session.rollback()
        except Exception:
            # ignore rollback errors
            pass

        # Step 1: Attempt legacy detailed PFASBOM query (may not exist under slim schema)
        bom_entries = []
        try:
            # Under slim schema PFASBOM has no per-material rows; this will usually be empty
            bom_entries = db.session.query(PFASBOM).filter_by(sku=sku).all()
        except Exception:
            # Ignore; we'll fallback
            bom_entries = []

        # If legacy BOM rows are not available, fallback to PPWRBOM materials for this SKU
        # This provides a minimal per-material list so assessment pages can still render.
        use_ppwr_bom_fallback = False
        try:
            if not bom_entries:
                ppwr_mats = db.session.query(PPWRBOM).filter_by(sku=sku).all()
                if not ppwr_mats:
                    return jsonify({"error": "SKU not found"}), 404
                use_ppwr_bom_fallback = True
        except Exception:
            pass

        # Derive product name for header
        if not use_ppwr_bom_fallback:
            first_row = bom_entries[0]
            product_name = f"{first_row.sku}_{first_row.product}"
        else:
            # Attempt to fetch product from slim PFASBOM summary row
            prod_val = None
            try:
                rec = db.session.query(PFASBOM).filter_by(sku=sku).first()
                if rec and getattr(rec, 'product', None):
                    prod_val = rec.product
            except Exception:
                pass
            product_name = f"{sku}_{prod_val}" if prod_val else f"{sku}_Unknown"

        # Step 2: Build query for materials + chemical and regulatory data
        if not use_ppwr_bom_fallback:
            # Legacy: expects detailed PFASBOM rows with material column
            try:
                material_ids = [entry.material for entry in bom_entries]
            except Exception:
                material_ids = []
            results = db.session.query(
                PFASBOM.component,
                PFASBOM.subcomponent,
                PFASBOM.material,
                PFASBOM.material_name,
                PFASMaterialChemicals.cas_number,
                PFASMaterialChemicals.chemical_name,
                PFASMaterialChemicals.concentration_ppm,
                PFASMaterialChemicals.supplier_name,
                PFASMaterialChemicals.reference_doc,
                PFASRegulations.australian_aics,
                PFASRegulations.australian_imap_tier_2,
                PFASRegulations.canadian_dsl,
                PFASRegulations.canada_pctsr_2012,
                PFASRegulations.eu_reach_pre_registered,
                PFASRegulations.eu_reach_registered_ppm,
                PFASRegulations.us_epa_tscainventory,
                PFASRegulations.us_epa_tsca12b
            ).outerjoin(PFASMaterialChemicals, PFASBOM.material == PFASMaterialChemicals.material_id) \
             .outerjoin(PFASRegulations, PFASMaterialChemicals.cas_number == PFASRegulations.cas_number) \
             .filter(PFASBOM.sku == sku).all()
        else:
            # Fallback: drive the table using PPWRBOM.material_id for this SKU
            results = db.session.query(
                PPWRBOM.material_id.label('material'),
                PFASMaterialChemicals.material_name,
                PFASMaterialChemicals.cas_number,
                PFASMaterialChemicals.chemical_name,
                PFASMaterialChemicals.concentration_ppm,
                PFASMaterialChemicals.supplier_name,
                PFASMaterialChemicals.reference_doc,
                PFASRegulations.australian_aics,
                PFASRegulations.australian_imap_tier_2,
                PFASRegulations.canadian_dsl,
                PFASRegulations.canada_pctsr_2012,
                PFASRegulations.eu_reach_pre_registered,
                PFASRegulations.eu_reach_registered_ppm,
                PFASRegulations.us_epa_tscainventory,
                PFASRegulations.us_epa_tsca12b
            ).outerjoin(PFASMaterialChemicals, PPWRBOM.material_id == PFASMaterialChemicals.material_id) \
             .outerjoin(PFASRegulations, PFASMaterialChemicals.cas_number == PFASRegulations.cas_number) \
             .filter(PPWRBOM.sku == sku).all()

        # Step 3: Define regulations
        regulations = [
            {'name': 'Australian AICS', 'col': 'australian_aics'},
            {'name': 'Australian IMAP Tier 2', 'col': 'australian_imap_tier_2'},
            {'name': 'Canadian DSL', 'col': 'canadian_dsl'},
            {'name': 'Canada PCTSR 2012', 'col': 'canada_pctsr_2012'},
            {'name': 'EU REACH Pre Registered', 'col': 'eu_reach_pre_registered'},
            {'name': 'EU REACH Registered', 'col': 'eu_reach_registered_ppm'},
            {'name': 'US EPA TSCA Inventory', 'col': 'us_epa_tscainventory'},
            {'name': 'US EPA TSCA 12B', 'col': 'us_epa_tsca12b'}
        ]

        data = []
        non_conforming_count = 0

        for row in results:
            # Use BOM data as fallback
            if use_ppwr_bom_fallback:
                component = "-"
                subcomponent = "-"
                material = getattr(row, 'material', None) or getattr(row, 'material_id', None) or "-"
                material_name = getattr(row, 'material_name', None) or "-"
            else:
                component = getattr(row, 'component', None) or "-"
                subcomponent = getattr(row, 'subcomponent', None) or "-"
                material = getattr(row, 'material', None) or "-"
                material_name = getattr(row, 'material_name', None) or "-"

            # Chemical data — use defaults if missing
            cas = row.cas_number or "Unknown"
            chem_name = row.chemical_name or "Unknown"

            # Handle concentration: if missing or invalid, set to None (will become "Unknown")
            try:
                conc = int(float(row.concentration_ppm) ) if row.concentration_ppm is not None else None
            except (ValueError, TypeError):
                conc = None

            supplier = row.supplier_name or "Unknown"

            # Check if this row has NO chemical data (i.e., left join produced nulls)
            has_chemical_data = row.cas_number is not None or row.chemical_name is not None

            limits = []
            is_non_conforming = False

            for reg in regulations:
                value = getattr(row, reg['col'], None)
                threshold = float(value) if value is not None else None

                if conc is None:
                    # 🚨 CRITICAL CHANGE: If concentration is unknown, mark as NON-CONFORMING for ALL regulations
                    status = 'exceeded'
                    color = 'danger'
                    limit_display = f"{threshold} ppm" if threshold is not None else "No Threshold"
                    is_non_conforming = True
                else:
                    # Only calculate compliance if we have a valid concentration
                    if threshold is None:
                        status = 'unknown'
                        color = 'warning'
                        limit_display = "No Data"
                    elif threshold < conc:
                        status = 'exceeded'
                        color = 'danger'
                        is_non_conforming = True
                        limit_display = f"{threshold} ppm"
                    else:
                        status = 'within'
                        color = 'success'
                        limit_display = f"{threshold} ppm"

                limits.append({
                    'name': reg['name'],
                    'limit': limit_display,
                    'status': status,
                    'color': color
                })

            if is_non_conforming:
                non_conforming_count += 1

            data.append({
                "component": component,
                "subcomponent": subcomponent,
                "material": material,
                "material_name": material_name,
                "supplier_name": supplier,
                "chemical_name": chem_name,
                "cas_number": cas,
                "concentration": f"{int(conc)} ppm" if conc is not None else "Unknown",
                "reference_doc": row.reference_doc or "—",
                "limits": limits,
                "status": "Non-Compliance" if is_non_conforming else ("No Chemical Data" if not has_chemical_data else "Compliance"),
                "status_color": "danger" if is_non_conforming else ("warning" if not has_chemical_data else "success")
            })

        in_conformance_count = len([d for d in data if d["status"] == "Compliance"])
        no_data_count = len([d for d in data if d["status"] == "No Chemical Data"])

        return jsonify({
            "product": product_name,
            "summary": {
                "total": len(data),
                "non_conforming": non_conforming_count,
                "in_conformance": in_conformance_count,
                "no_chemical_data": no_data_count
            },
            "data": data
        })

    except Exception as e:
        app.logger.error(f"Error in pfas_assessment for {sku}: {e}", exc_info=True)
        return jsonify({"error": "Failed to retrieve data"}), 500

def calculate_dynamic_summary(sku, assessment_data, strict: bool = False):
    """Calculate dynamic summary statistics based on regulatory thresholds.
    
    Checks each chemical against all regulatory limits from pfas_regulation table.
    Uses strict mode for PPWR (unknown concentration = non-conforming) or legacy
    mode for PFAS (unknown concentration = no data).
    
    Args:
        sku: Product SKU to analyze
        assessment_data: Dict containing assessment data with 'data' key
        strict: If True, treat unknown concentrations as non-conforming (PPWR mode)
    
    Returns:
        Dict with keys:
            files: {total, downloaded, not_found, progress_text}
            review: {reviewed, total_expected, non_conforming, in_conformance,
                    no_data, alt_found, alt_not_found, progress_text}
    """
    try:
        app.logger.info(f"🔍 Starting regulatory-based summary calculation for SKU: {sku}")
        
        # Get assessment data
        data_entries = assessment_data.get('data', [])
        app.logger.info(f"📋 Processing {len(data_entries)} assessment entries")
        
        # Calculate file progress metrics
        total_materials_query = db.session.query(PFASBOM.material).filter_by(sku=sku).distinct()
        total_materials = total_materials_query.count()
        
        # Count materials that have chemical data
        materials_with_chemicals_query = db.session.query(PFASBOM.material).filter_by(sku=sku).join(
            PFASMaterialChemicals, PFASBOM.material == PFASMaterialChemicals.material_id
        ).distinct()
        materials_with_chemicals = materials_with_chemicals_query.count()
        
        files_not_found = max(0, total_materials - materials_with_chemicals)
        
        app.logger.info(f"📁 Materials: Total={total_materials}, With Chemicals={materials_with_chemicals}, Not Found={files_not_found}")
        
        # Calculate conformance based on actual regulatory checks.
        # In strict mode (PPWR) unknown concentrations count as non-conforming.
        conformance_stats = calculate_regulatory_conformance(sku, strict=strict)

        # Calculate review metrics
        total_reviewed = len(data_entries)
        estimated_total = max(total_reviewed, total_materials) if total_reviewed > 0 else total_materials

        # Calculate alternative substance metrics based on conformance
        alt_found = max(0, conformance_stats['in_conformance'] // 3) if conformance_stats['in_conformance'] > 0 else 0
        alt_not_found = max(0, conformance_stats['non_conforming'] // 4) if conformance_stats['non_conforming'] > 0 else 0

        summary_stats = {
            'files': {
                'total': total_materials,
                'downloaded': materials_with_chemicals,
                'not_found': files_not_found,
                'progress_text': f"{materials_with_chemicals} / {total_materials}"
            },
            'review': {
                'reviewed': total_reviewed,
                'total_expected': estimated_total,
                'non_conforming': conformance_stats['non_conforming'],
                'in_conformance': conformance_stats['in_conformance'],
                'no_data': conformance_stats['no_chemical_data'],
                'alt_found': alt_found,
                'alt_not_found': alt_not_found,
                'progress_text': f"{total_reviewed} / {estimated_total}"
            }
        }

        app.logger.info(f"✅ Final regulatory-based summary stats: {summary_stats}")
        return summary_stats

    except Exception as e:
        app.logger.error(f"❌ Error calculating regulatory summary for {sku}: {e}", exc_info=True)
        return {
            'files': {
                'total': 0,
                'downloaded': 0,
                'not_found': 0,
                'progress_text': "Error / Error"
            },
            'review': {
                'reviewed': 0,
                'total_expected': 0,
                'non_conforming': 0,
                'in_conformance': 0,
                'no_data': 0,
                'alt_found': 0,
                'alt_not_found': 0,
                'progress_text': "Error / Error"
            }
        }


def calculate_regulatory_conformance(sku, strict: bool = False):
    """Calculate conformance/non-conformance based on actual regulatory thresholds.
    
    Evaluates materials against all applicable regulations (Australian AICS, IMAP,
    Canadian DSL, PCTSR, EU REACH, US EPA TSCA) and categorizes each as conforming,
    non-conforming, or no data available.
    
    Args:
        sku: Product SKU to analyze
        strict: If True (PPWR mode), unknown concentration = non-conforming.
                If False (PFAS mode), unknown concentration = no_chemical_data
    
    Returns:
        dict: {
            'non_conforming': int,  # Materials exceeding any threshold
            'in_conformance': int,  # Materials within all thresholds
            'no_chemical_data': int # Materials with missing data
        }
    """
    try:
        app.logger.info(f"🏛️ Calculating regulatory conformance for SKU: {sku}")
        
        # Define regulatory columns to check
        regulatory_columns = [
            'australian_aics',
            'australian_imap_tier_2', 
            'canadian_dsl',
            'canada_pctsr_2012',
            'eu_reach_pre_registered',
            'eu_reach_registered_ppm',
            'us_epa_tscainventory',
            'us_epa_tsca12b'
        ]
        
        # Query to get all materials with their chemicals and regulatory data
        results = db.session.query(
            PFASBOM.material,
            PFASBOM.material_name,
            PFASMaterialChemicals.cas_number,
            PFASMaterialChemicals.chemical_name,
            PFASMaterialChemicals.concentration_ppm,
            PFASRegulations.australian_aics,
            PFASRegulations.australian_imap_tier_2,
            PFASRegulations.canadian_dsl,
            PFASRegulations.canada_pctsr_2012,
            PFASRegulations.eu_reach_pre_registered,
            PFASRegulations.eu_reach_registered_ppm,
            PFASRegulations.us_epa_tscainventory,
            PFASRegulations.us_epa_tsca12b
        ).outerjoin(
            PFASMaterialChemicals, PFASBOM.material == PFASMaterialChemicals.material_id
        ).outerjoin(
            PFASRegulations, PFASMaterialChemicals.cas_number == PFASRegulations.cas_number
        ).filter(PFASBOM.sku == sku).all()
        
        app.logger.info(f"📊 Found {len(results)} material-chemical-regulation records")
        
        # Counters
        non_conforming = 0
        in_conformance = 0
        no_chemical_data = 0
        
        # Track processed entries to avoid double counting
        processed_entries = set()
        
        for row in results:
            # Create unique identifier for this material-chemical combination
            entry_id = f"{row.material}_{row.cas_number}_{row.chemical_name}"
            
            if entry_id in processed_entries:
                continue  # Skip duplicates
            processed_entries.add(entry_id)
            
            # Check if this entry has chemical data
            has_chemical_data = (row.cas_number is not None and 
                               row.chemical_name is not None and
                               row.concentration_ppm is not None)
            
            if not has_chemical_data:
                no_chemical_data += 1
                app.logger.debug(f"📝 No chemical data for material {row.material}")
                continue
            
            # Parse concentration
            try:
                concentration = float(row.concentration_ppm) if row.concentration_ppm else 0.0
            except (ValueError, TypeError):
                concentration = 0.0
                app.logger.warning(f"⚠️ Invalid concentration for {row.cas_number}: {row.concentration_ppm}")
            
            # Check conformance against all regulatory thresholds
            is_non_conforming = False
            conformance_checks = 0
            
            for reg_col in regulatory_columns:
                threshold_value = getattr(row, reg_col, None)
                
                if threshold_value is not None:
                    conformance_checks += 1
                    threshold = float(threshold_value)
                    
                    # If concentration is unknown (0 or None): behavior depends on 'strict'
                    if concentration == 0.0 or concentration is None:
                        if strict:
                            is_non_conforming = True
                            app.logger.debug(f"❌ Non-conforming due to unknown concentration (strict mode): {row.chemical_name} ({row.cas_number})")
                            break
                        else:
                            # In legacy (non-strict) mode, treat as missing chemical data (counted elsewhere)
                            app.logger.debug(f"ℹ️ Unknown concentration treated as no_chemical_data (non-strict mode): {row.chemical_name} ({row.cas_number})")
                            # Stop checking further regulations for this entry
                            break
                    elif concentration > threshold:
                        is_non_conforming = True
                        app.logger.debug(f"❌ Non-conforming: {row.chemical_name} ({concentration} ppm > {threshold} ppm for {reg_col})")
                        break
            
            # Categorize the entry
            if is_non_conforming:
                non_conforming += 1
            elif conformance_checks > 0:  # Has regulatory data and passed all checks
                in_conformance += 1
                app.logger.debug(f"✅ Conforming: {row.chemical_name} passed all {conformance_checks} regulatory checks")
            else:
                # Has chemical data but no regulatory thresholds to compare against
                no_chemical_data += 1
                app.logger.debug(f"📝 No regulatory data for {row.chemical_name} ({row.cas_number})")
        
        # Handle materials with no chemical data at all
        materials_without_chemicals = db.session.query(PFASBOM.material).filter_by(sku=sku).filter(
            ~db.session.query(PFASMaterialChemicals.material_id).filter(
                PFASMaterialChemicals.material_id == PFASBOM.material
            ).exists()
        ).count()
        
        no_chemical_data += materials_without_chemicals
        
        conformance_result = {
            'non_conforming': non_conforming,
            'in_conformance': in_conformance, 
            'no_chemical_data': no_chemical_data
        }
        
        app.logger.info(f"🏛️ Regulatory conformance results: {conformance_result}")
        return conformance_result
        
    except Exception as e:
        app.logger.error(f"❌ Error calculating regulatory conformance: {e}", exc_info=True)
        return {
            'non_conforming': 0,
            'in_conformance': 0,
            'no_chemical_data': 0
        }
    
@app.route('/download-bom/<sku>')
def download_bom(sku):
    """Deprecated legacy route. Redirect to the new download endpoint.

    Keeping this for backward compatibility; logs usage and issues a 301 to
    /bom/download/<sku> implemented by api_bom_download_bytes.
    """
    try:
        app.logger.info(f"DEPRECATED: /download-bom/{sku} called, redirecting to /bom/download/{sku}")
    except Exception:
        pass
    return redirect(url_for('api_bom_download_bytes', sku=sku), code=301)

@app.route('/bom/download/<sku>')
def api_bom_download_bytes(sku):
    """Reconstruct and download BOM CSV from ppwr_bom rows for this SKU."""
    try:
        # Query all materials for this SKU from ppwr_bom
        materials = db.session.query(PPWRBOM).filter_by(sku=sku).all()
        
        if not materials:
            flash("No BOM data found for this SKU.", "danger")
            return redirect(url_for('index'))
        
        # Build CSV with headers: SKU, Material ID, Material Name, Supplier, Uploaded At
        csv_lines = ["SKU,Material ID,Material Name,Supplier,Uploaded At"]
        
        for mat in materials:
            mat_id = mat.material_id or ''
            mat_name = (mat.material_name or '').replace(',', ';')  # Escape commas
            supplier = (mat.supplier_name or '').replace(',', ';')
            uploaded = mat.uploaded_at.strftime('%Y-%m-%d %H:%M:%S') if mat.uploaded_at else ''
            
            csv_lines.append(f"{sku},{mat_id},{mat_name},{supplier},{uploaded}")
        
        csv_content = "\n".join(csv_lines)
        csv_bytes = csv_content.encode('utf-8')
        
        download_name = f"BOM_{sku}.csv"
        return send_file(
            BytesIO(csv_bytes),
            as_attachment=True,
            download_name=download_name,
            mimetype='text/csv'
        )
    except Exception as e:
        app.logger.error(f"BOM download failed for {sku}: {e}", exc_info=True)
        flash("Error downloading BOM.", "danger")
        return redirect(url_for('index'))


@app.route('/api/delete-product/<sku>', methods=['DELETE'])
def delete_product(sku):
    """Delete a product (route entry + ppwr_bom materials)"""
    app.logger.info(f"Attempting to delete product with SKU: {sku}")
    
    try:
        # Delete route entry
        route_deleted = db.session.query(Route).filter_by(sku=sku).delete()
        
        # Delete all ppwr_bom entries for this SKU
        materials_deleted = db.session.query(PPWRBOM).filter_by(sku=sku).delete()
        
        if route_deleted == 0 and materials_deleted == 0:
            app.logger.warning(f"No entries found for SKU: {sku}")
            return jsonify({"error": "Product not found"}), 404
        
        # Commit the deletion
        db.session.commit()
        
        app.logger.info(f"Successfully deleted SKU {sku}: {route_deleted} route(s), {materials_deleted} material(s)")
        
        return jsonify({
            "success": True,
            "message": f"Product '{sku}' deleted successfully",
            "route_deleted": route_deleted,
            "materials_deleted": materials_deleted,
            "sku": sku
        }), 200
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error deleting product {sku}: {e}", exc_info=True)
        return jsonify({
            "error": "Failed to delete product. Please try again.",
            "details": str(e)
        }), 500

   
@app.route('/api/export-pfas-report/<sku>')
def export_pfas_report(sku):
    """
    Exports a comprehensive PFAS report for a single SKU.
    Includes BOM, chemical details, and regulatory limits with compliance status.
    """
    app.logger.info(f"📥 Generating PFAS report for SKU: {sku}")

    try:
        # Fetch BOM entries for the SKU
        bom_entries = db.session.query(PFASBOM).filter_by(sku=sku).all()
        if not bom_entries:
            return jsonify({"error": "SKU not found"}), 404

        product_name = bom_entries[0].product

        # Regulation columns
        regulation_columns = [
            'australian_aics',
            'australian_imap_tier_2',
            'canadian_dsl',
            'canada_pctsr_2012',
            'eu_reach_pre_registered',
            'eu_reach_registered_ppm',
            'us_epa_tscainventory',
            'us_epa_tsca12b'
        ]

        # Build the full query
        query = db.session.query(
            PFASBOM.sku,
            PFASBOM.product,
            PFASBOM.component,
            PFASBOM.component_description,
            PFASBOM.subcomponent,
            PFASBOM.subcomponent_description,
            PFASBOM.material,
            PFASBOM.material_name,
            PFASMaterialChemicals.cas_number,
            PFASMaterialChemicals.chemical_name,
            PFASMaterialChemicals.concentration_ppm,
            PFASMaterialChemicals.supplier_name,
            PFASMaterialChemicals.reference,
            PFASRegulations.australian_aics,
            PFASRegulations.australian_imap_tier_2,
            PFASRegulations.canadian_dsl,
            PFASRegulations.canada_pctsr_2012,
            PFASRegulations.eu_reach_pre_registered,
            PFASRegulations.eu_reach_registered_ppm,
            PFASRegulations.us_epa_tscainventory,
            PFASRegulations.us_epa_tsca12b
        ).join(PFASMaterialChemicals, PFASBOM.material == PFASMaterialChemicals.material_id)\
         .outerjoin(PFASRegulations, PFASMaterialChemicals.cas_number == PFASRegulations.cas_number)\
         .filter(PFASBOM.sku == sku)

        results = query.all()
        app.logger.info(f"📊 PFAS report query returned {len(results)} rows")

        if not results:
            return jsonify({"error": "No chemical data found for this SKU"}), 400

        # Prepare data for DataFrame
        rows = []
        for r in results:
            conc = int(float(r.concentration_ppm)) if r.concentration_ppm is not None else 0.0

            row = {
                "SKU": r.sku or "",
                "Product": r.product or "",
                "Component": r.component or "",
                "Component Description": r.component_description or "",
                "Sub-Component": r.subcomponent or "",
                "Sub-Component Description": r.subcomponent_description or "",
                "Material ID": r.material or "",
                "Material Name": r.material_name or "",
                "CAS Number": r.cas_number or "Unknown",
                "Chemical Name": r.chemical_name or "Unknown",
                "Chemical Concentration (ppm)": f"{int(conc)} ppm",
                "Supplier Name": r.supplier_name or "Unknown",
                "Reference": r.reference or ""
            }

            # Add regulation thresholds and compliance status
            for col in regulation_columns:
                value = getattr(r, col, None)
                if value is None:
                    row[col] = "No Data"
                else:
                    threshold = float(value)
                    status = "Non-Compliant" if conc > threshold else "Compliant"
                    row[col] = {
                        "value": threshold,
                        "status": "exceeded" if conc > threshold else "within"
                    }

            rows.append(row)

        # Create DataFrame
        df_data = []
        for row in rows:
            flat_row = {k: (v["value"] if isinstance(v, dict) else v) for k, v in row.items()}
            # Format ppm values
            if "Chemical Concentration (ppm)" not in flat_row:
                conc_val = row.get("Chemical Concentration (ppm)", "0.00 ppm")
                flat_row["Chemical Concentration (ppm)"] = conc_val
            df_data.append(flat_row)

        df = pd.DataFrame(df_data)

        # Write to Excel with styling
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='PFAS Assessment', index=False)
            worksheet = writer.sheets['PFAS Assessment']

            # Styling
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green

            # Apply coloring for regulation columns
            start_row = 2
            col_idx_map = {col: df.columns.get_loc(col) + 1 for col in regulation_columns}
            for r_idx, row in enumerate(rows):
                for reg_col in regulation_columns:
                    if reg_col not in col_idx_map:
                        continue
                    cell = worksheet.cell(row=start_row + r_idx, column=col_idx_map[reg_col])
                    value_obj = row[reg_col]
                    if isinstance(value_obj, dict):
                        if value_obj["status"] == "exceeded":
                            cell.fill = red_fill
                        else:
                            cell.fill = green_fill

            # Auto-adjust column widths
            for i, col in enumerate(df.columns):
                max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.column_dimensions[chr(65 + i)].width = min(max_len, 50)

        output.seek(0)
        timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M")
        filename = f"PFAS_Report_{sku}_{timestamp}.xlsx"

        app.logger.info(f"✅ PFAS report generated: {filename}")

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        app.logger.error(f"❌ Failed to generate PFAS report for {sku}: {e}", exc_info=True)
        return jsonify({"error": "Failed to generate report. Please try again."}), 500


@app.route("/download/pfas/<sku>")
def download_pfas_report(sku):
    try:
        app.logger.info(f"📥 Generating PFAS report for SKU={sku}")

        # Regulation columns to include
        regulation_columns = [
            #'australian_aics',
            #'australian_imap_tier_2',
            #'canadian_dsl',
            'canada_pctsr_2012',
            'eu_reach_pre_registered',
            'eu_reach_registered_ppm',
            'us_epa_tscainventory',
            'us_epa_tsca12b'
        ]

        # Build query: Join BOM → Chemicals → Regulations
        query = db.session.query(
            PFASBOM.sku,
            PFASBOM.product,
            PFASBOM.component,
            PFASBOM.component_description,
            PFASBOM.subcomponent,
            PFASBOM.subcomponent_description,
            PFASBOM.material,
            PFASBOM.material_name,
            PFASMaterialChemicals.cas_number,
            PFASMaterialChemicals.chemical_name,
            PFASMaterialChemicals.concentration_ppm,
            PFASMaterialChemicals.supplier_name,
            PFASMaterialChemicals.reference_doc,
            # Regulation thresholds
            PFASRegulations.australian_aics,
            PFASRegulations.australian_imap_tier_2,
            PFASRegulations.canadian_dsl,
            PFASRegulations.canada_pctsr_2012,
            PFASRegulations.eu_reach_pre_registered,
            PFASRegulations.eu_reach_registered_ppm,
            PFASRegulations.us_epa_tscainventory,
            PFASRegulations.us_epa_tsca12b
        ).join(PFASMaterialChemicals, PFASBOM.material == PFASMaterialChemicals.material_id) \
         .outerjoin(PFASRegulations, PFASMaterialChemicals.cas_number == PFASRegulations.cas_number) \
         .filter(PFASBOM.sku == sku)

        results = query.all()
        app.logger.info(f"📊 Query returned {len(results)} rows")

        if not results:
            return jsonify({"error": "No matching records found"}), 400

        # Prepare rows
        rows = []
        for r in results:
            conc = int(float(r.concentration_ppm)) if r.concentration_ppm is not None else 0.0

            row = {
                "SKU": r.sku or "",
                "Product": r.product or "",
                "Component": r.component or "",
                "Component Description": r.component_description or "",
                "Sub-Component": r.subcomponent or "",
                "Sub-Component Description": r.subcomponent_description or "",
                "Material ID": r.material or "",
                "Material Name": r.material_name or "",
                "CAS Number": r.cas_number or "Unknown",
                "Chemical Name": r.chemical_name or "Unknown",
                "Chemical Concentration": f"{int(conc)} ppm",
                "Supplier Name": r.supplier_name or "Unknown",
                "Reference": r.reference_doc or ""
            }

            # Add regulation values
            for col in regulation_columns:
                value = getattr(r, col, None)
                if value is None:
                    row[col] = "No Data"
                else:
                    threshold = float(value)
                    row[col] = {
                        "value": threshold,
                        "status": "exceeded" if conc > threshold else "within"
                    }

            rows.append(row)

        # Flatten for DataFrame
        df_data = []
        for row in rows:
            flat = {k: v for k, v in row.items() if not isinstance(v, dict)}
            for col in regulation_columns:
                if isinstance(row[col], dict):
                    flat[col] = f"{row[col]['value']} ppm"
                else:
                    flat[col] = row[col]
            df_data.append(flat)

        df = pd.DataFrame(df_data)

        # Write to Excel with coloring
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='PFAS Report', index=False)
            worksheet = writer.sheets['PFAS Report']

            # Coloring
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

            start_row = 2
            for idx, row in enumerate(rows):
                for col_idx, col_name in enumerate(regulation_columns):
                    cell = worksheet.cell(row=start_row + idx, column=df.columns.get_loc(col_name) + 1)
                    value_obj = row[col_name]
                    if isinstance(value_obj, dict):
                        if value_obj["status"] == "exceeded":
                            cell.fill = red_fill
                        else:
                            cell.fill = green_fill

            # Auto-fit column width
            for i, col in enumerate(df.columns):
                max_len = max(df[col].astype(str).map(len).max(), len(col))
                worksheet.column_dimensions[chr(65 + i)].width = min(max_len + 2, 50)

        output.seek(0)
        now = pd.Timestamp.now().strftime("%Y%m%d_%H%M")
        filename = f"PFAS_Report_{sku}_{now}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        app.logger.error(f"❌ PFAS report export failed: {e}", exc_info=True)
        return jsonify({"error": "Export failed. Please try again."}), 500

# Add this test route to debug the summary calculation
@app.route('/test-summary/<sku>')
def test_summary(sku):
    """Test route to check summary calculation without full assessment"""
    try:
        app.logger.info(f"🧪 Testing summary calculation for SKU: {sku}")
        
        # Check BOM data
        bom_count = db.session.query(PFASBOM).filter_by(sku=sku).count()
        bom_materials = db.session.query(PFASBOM.material).filter_by(sku=sku).distinct().count()
        
        # Check chemical data
        chemical_count = db.session.query(PFASMaterialChemicals).join(
            PFASBOM, PFASMaterialChemicals.material_id == PFASBOM.material
        ).filter(PFASBOM.sku == sku).count()
        
        materials_with_chemicals = db.session.query(PFASBOM.material).filter_by(sku=sku).join(
            PFASMaterialChemicals, PFASBOM.material == PFASMaterialChemicals.material_id
        ).distinct().count()
        
        # Try to get PFAS assessment
        try:
            response = pfas_assessment(sku)
            pfas_data = response.get_json() if response.status_code == 200 else {"error": "Assessment failed"}
        except Exception as e:
            pfas_data = {"error": str(e)}
        
        # Try summary calculation
        try:
            if "error" not in pfas_data:
                # Use legacy (non-strict) behavior for PFAS summary by default
                summary_stats = calculate_dynamic_summary(sku, pfas_data, strict=False)
            else:
                summary_stats = {"error": "Could not calculate due to PFAS assessment failure"}
        except Exception as e:
            summary_stats = {"error": str(e)}
        
        return jsonify({
            "sku": sku,
            "bom_data": {
                "total_bom_records": bom_count,
                "unique_materials": bom_materials
            },
            "chemical_data": {
                "total_chemical_records": chemical_count,
                "materials_with_chemicals": materials_with_chemicals
            },
            "pfas_assessment": {
                "status": "success" if "error" not in pfas_data else "failed",
                "data_count": len(pfas_data.get('data', [])) if "error" not in pfas_data else 0,
                "summary": pfas_data.get('summary', {}) if "error" not in pfas_data else pfas_data
            },
            "calculated_summary": summary_stats
        })
        
    except Exception as e:
        app.logger.error(f"❌ Test summary failed for {sku}: {e}", exc_info=True)
        return jsonify({"error": str(e), "sku": sku})
    

# Enhanced test route to debug regulatory calculations
@app.route('/test-regulatory/<sku>')
def test_regulatory(sku):
    """Test route to check regulatory-based conformance calculation"""
    try:
        app.logger.info(f"🧪 Testing regulatory calculations for SKU: {sku}")
        
        # Check basic data counts
        bom_count = db.session.query(PFASBOM).filter_by(sku=sku).count()
        unique_materials = db.session.query(PFASBOM.material).filter_by(sku=sku).distinct().count()
        
        # Check chemical data
        chemicals_query = db.session.query(
            PFASMaterialChemicals.material_id,
            PFASMaterialChemicals.cas_number, 
            PFASMaterialChemicals.chemical_name,
            PFASMaterialChemicals.concentration_ppm
        ).join(PFASBOM, PFASMaterialChemicals.material_id == PFASBOM.material).filter(PFASBOM.sku == sku)
        
        chemicals = chemicals_query.all()
        
        # Check regulatory data
        regulatory_query = db.session.query(
            PFASRegulations.cas_number,
            PFASRegulations.australian_aics,
            PFASRegulations.australian_imap_tier_2,
            PFASRegulations.canadian_dsl,
            PFASRegulations.canada_pctsr_2012,
            PFASRegulations.eu_reach_pre_registered,
            PFASRegulations.eu_reach_registered_ppm,
            PFASRegulations.us_epa_tscainventory,
            PFASRegulations.us_epa_tsca12b
        ).join(PFASMaterialChemicals, PFASRegulations.cas_number == PFASMaterialChemicals.cas_number).join(
            PFASBOM, PFASMaterialChemicals.material_id == PFASBOM.material
        ).filter(PFASBOM.sku == sku)
        
        regulatory_data = regulatory_query.all()
        
        # Test the conformance calculation
        try:
            # Legacy/test route should call non-strict mode to reflect PFAS legacy behaviour
            conformance_stats = calculate_regulatory_conformance(sku, strict=False)
        except Exception as e:
            conformance_stats = {"error": str(e)}
        
        # Sample some data for inspection
        sample_chemicals = []
        for chem in chemicals[:5]:  # First 5 chemicals
            sample_chemicals.append({
                "material_id": chem.material_id,
                "cas_number": chem.cas_number,
                "chemical_name": chem.chemical_name,
                "concentration_ppm": chem.concentration_ppm
            })
        
        sample_regulatory = []
        for reg in regulatory_data[:5]:  # First 5 regulatory entries
            sample_regulatory.append({
                "cas_number": reg.cas_number,
                "australian_aics": reg.australian_aics,
                "australian_imap_tier_2": reg.australian_imap_tier_2,
                "canadian_dsl": reg.canadian_dsl,
                "canada_pctsr_2012": reg.canada_pctsr_2012,
                "eu_reach_pre_registered": reg.eu_reach_pre_registered,
                "eu_reach_registered_ppm": reg.eu_reach_registered_ppm,
                "us_epa_tscainventory": reg.us_epa_tscainventory,
                "us_epa_tsca12b": reg.us_epa_tsca12b
            })
        
        return jsonify({
            "sku": sku,
            "basic_counts": {
                "bom_records": bom_count,
                "unique_materials": unique_materials,
                "chemical_records": len(chemicals),
                "regulatory_records": len(regulatory_data)
            },
            "sample_data": {
                "chemicals": sample_chemicals,
                "regulatory": sample_regulatory
            },
            "conformance_calculation": conformance_stats,
            "summary_calculation": {
                "status": "attempting...",
                "result": "Check logs for detailed calculation steps"
            }
        })
        
    except Exception as e:
        app.logger.error(f"❌ Test regulatory failed for {sku}: {e}", exc_info=True)
        return jsonify({"error": str(e), "sku": sku})


# Also add a simplified debug route to see raw query results
@app.route('/debug-raw-data/<sku>')
def debug_raw_data(sku):
    """Debug route to see raw data from the main regulatory query"""
    try:
        # The exact query used in calculate_regulatory_conformance
        results = db.session.query(
            PFASBOM.material,
            PFASBOM.material_name,
            PFASMaterialChemicals.cas_number,
            PFASMaterialChemicals.chemical_name,
            PFASMaterialChemicals.concentration_ppm,
            PFASRegulations.australian_aics,
            PFASRegulations.australian_imap_tier_2,
            PFASRegulations.canadian_dsl,
            PFASRegulations.canada_pctsr_2012,
            PFASRegulations.eu_reach_pre_registered,
            PFASRegulations.eu_reach_registered_ppm,
            PFASRegulations.us_epa_tscainventory,
            PFASRegulations.us_epa_tsca12b
        ).outerjoin(
            PFASMaterialChemicals, PFASBOM.material == PFASMaterialChemicals.material_id
        ).outerjoin(
            PFASRegulations, PFASMaterialChemicals.cas_number == PFASRegulations.cas_number
        ).filter(PFASBOM.sku == sku).limit(10).all()  # Limit to first 10 for readability
        
        formatted_results = []
        for row in results:
            formatted_results.append({
                "material": row.material,
                "material_name": row.material_name,
                "cas_number": row.cas_number,
                "chemical_name": row.chemical_name,
                "concentration_ppm": row.concentration_ppm,
                "regulatory_thresholds": {
                    "australian_aics": row.australian_aics,
                    "australian_imap_tier_2": row.australian_imap_tier_2,
                    "canadian_dsl": row.canadian_dsl,
                    "canada_pctsr_2012": row.canada_pctsr_2012,
                    "eu_reach_pre_registered": row.eu_reach_pre_registered,
                    "eu_reach_registered_ppm": row.eu_reach_registered_ppm,
                    "us_epa_tscainventory": row.us_epa_tscainventory,
                    "us_epa_tsca12b": row.us_epa_tsca12b
                }
            })
        
        return jsonify({
            "sku": sku,
            "query_results_sample": formatted_results,
            "total_query_results": db.session.query(
                PFASBOM.material
            ).outerjoin(
                PFASMaterialChemicals, PFASBOM.material == PFASMaterialChemicals.material_id
            ).outerjoin(
                PFASRegulations, PFASMaterialChemicals.cas_number == PFASRegulations.cas_number
            ).filter(PFASBOM.sku == sku).count()
        })
        
    except Exception as e:
        return jsonify({"error": str(e), "sku": sku})
    
# ==================== LEGACY FILTER ROUTES REMOVED ====================
# Deprecated routes removed: /api/skus, /api/components, /api/subcomponents, /api/materials
# Use unified assessment pages with search functionality instead

@app.route('/api/export-excel/<sku>')
def export_excel(sku):
    try:
        # Fetch assessment data by calling the internal function and extracting JSON
        resp = pfas_assessment(sku)
        if hasattr(resp, 'status_code') and resp.status_code != 200:
            app.logger.error(f"export_excel: pfas_assessment failed for {sku}: {resp.get_data(as_text=True)}")
            return jsonify({'error': 'Assessment failed'}), 400

        assessment_json = resp.get_json() if hasattr(resp, 'get_json') else resp
        data = assessment_json.get('data', []) if isinstance(assessment_json, dict) else []

        # Create a DataFrame
        df = pd.DataFrame(data)

        # Create an Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='PFAS Assessment')

        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'PFAS-Assessment-{sku}.xlsx'
        )
    except Exception as e:
        app.logger.error(f"Error exporting to Excel: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500
    
# REMOVED: /api/filter-results - deprecated legacy route
# REMOVED: /api/export-filter-results - deprecated legacy route

# REMOVED: Deprecated incomplete duplicate route from line 3222 (kept complete version below)

# Add these new routes to your app.py (or replace the existing ones if duplicated)
@app.route('/api/assessment-regions/<sku>')
def get_assessment_regions(sku):
    """Get available regions for the assessment page"""
    try:
        # Define region mappings - corrected to match database columns and frontend keys
        region_mappings = {
            'australia_aics': 'Australia (AICS)',
            'australia_imap_tier_2': 'Australia (IMAP Tier 2)', # Corrected key
            'canadian_dsl': 'Canada (DSL)',
            'canada_pctsr_2012': 'Canada (PCTSR 2012)', # Corrected key
            'eu_reach_pre_registered': 'EU (REACH Pre-registered)', # Corrected key
            'eu_reach_registered_ppm': 'EU (REACH Registered)', # Corrected key
            'us_epa_tscainventory': 'USA (EPA TSCA Inventory)', # Corrected key
            'us_epa_tsca12b': 'USA (EPA TSCA 12B)' # Corrected key
        }

        # Column to frontend key mapping (database column name -> region_mappings key)
        column_to_region_key = {
            'australian_aics': 'australia_aics',
            'australian_imap_tier_2': 'australia_imap_tier_2', # Corrected
            'canadian_dsl': 'canadian_dsl',
            'canada_pctsr_2012': 'canada_pctsr_2012', # Corrected
            'eu_reach_pre_registered': 'eu_reach_pre_registered', # Corrected
            'eu_reach_registered_ppm': 'eu_reach_registered_ppm', # Corrected
            'us_epa_tscainventory': 'us_epa_tscainventory', # Corrected
            'us_epa_tsca12b': 'us_epa_tsca12b' # Corrected
        }

        # Check which regulatory columns have data for this SKU
        available_regions = []

        # Query to check which regulatory columns have non-null values for materials in this SKU
        # Use distinct to avoid checking the same regulation row multiple times if a chemical is used in multiple BOM entries
        # Select distinct cas_numbers relevant to this SKU first
        relevant_cas_numbers = db.session.query(PFASMaterialChemicals.cas_number).\
            join(PFASBOM, PFASMaterialChemicals.material_id == PFASBOM.material).\
            filter(PFASBOM.sku == sku).\
            distinct()

        # Instead of pulling all regulation rows and scanning them (which can be
        # fragile depending on how SQLAlchemy returns tuples), explicitly check
        # for each regulation column whether any PFASRegulations entry exists
        # (for CAS numbers present in this SKU) with a non-null value. This
        # is more robust and will correctly surface regions that have data.
        found_regions = set()
        for db_col, region_key in column_to_region_key.items():
            try:
                col_attr = getattr(PFASRegulations, db_col)
            except AttributeError:
                # Skip if mapping is wrong
                continue

            exists_q = db.session.query(PFASRegulations).filter(
                PFASRegulations.cas_number.in_(relevant_cas_numbers.subquery()),
                col_attr.isnot(None)
            ).limit(1).count()

            if exists_q and region_key in region_mappings:
                found_regions.add(region_key)

        # Convert to the format expected by frontend
        for region_key in sorted(found_regions):
            available_regions.append({
                'value': region_key,
                'label': region_mappings.get(region_key, region_key)
            })

        app.logger.info(f"Available regions for SKU {sku}: {available_regions}")
        return jsonify({
            'regions': available_regions,
            'default': 'all'
        })

    except Exception as e:
        app.logger.error(f"Error getting regions for {sku}: {e}", exc_info=True)
        return jsonify({'regions': [], 'default': 'all'}), 500


@app.route('/api/assessment-filter/<sku>')
def filter_assessment_by_region(sku):
    """Filter assessment data by region"""
    try:
        region_filter = request.args.get('region', 'all')

        app.logger.info(f"Filtering assessment for SKU {sku} by region: {region_filter}")

        # Define column mappings for regions - corrected keys to match database columns
        region_to_columns = {
            'australia_aics': 'australian_aics', # Map frontend key to actual DB column
            'australia_imap_tier_2': 'australian_imap_tier_2', # Corrected
            'canadian_dsl': 'canadian_dsl', # Corrected
            'canada_pctsr_2012': 'canada_pctsr_2012', # Corrected
            'eu_reach_pre_registered': 'eu_reach_pre_registered', # Corrected
            'eu_reach_registered_ppm': 'eu_reach_registered_ppm', # Corrected
            'us_epa_tscainventory': 'us_epa_tscainventory', # Corrected
            'us_epa_tsca12b': 'us_epa_tsca12b' # Corrected
        }

        # Define regulations for display and processing - ensure keys match region_to_columns keys
        regulations_definitions = [
            {'name': 'Australian AICS', 'col': 'australian_aics', 'region_key': 'australia_aics'},
            {'name': 'Australian IMAP Tier 2', 'col': 'australian_imap_tier_2', 'region_key': 'australia_imap_tier_2'}, # Corrected
            {'name': 'Canadian DSL', 'col': 'canadian_dsl', 'region_key': 'canadian_dsl'}, # Corrected
            {'name': 'Canada PCTSR 2012', 'col': 'canada_pctsr_2012', 'region_key': 'canada_pctsr_2012'}, # Corrected
            {'name': 'EU REACH Pre Registered', 'col': 'eu_reach_pre_registered', 'region_key': 'eu_reach_pre_registered'}, # Corrected
            {'name': 'EU REACH Registered', 'col': 'eu_reach_registered_ppm', 'region_key': 'eu_reach_registered_ppm'}, # Corrected
            {'name': 'US EPA TSCA Inventory', 'col': 'us_epa_tscainventory', 'region_key': 'us_epa_tscainventory'}, # Corrected
            {'name': 'US EPA TSCA 12B', 'col': 'us_epa_tsca12b', 'region_key': 'us_epa_tsca12b'} # Corrected
        ]

        # Filter regulations based on selected region
        filtered_regulations = regulations_definitions
        if region_filter != 'all' and region_filter in region_to_columns:
             # Filter regulations to only include the one selected
            filtered_regulations = [reg for reg in regulations_definitions if reg['region_key'] == region_filter]
            app.logger.debug(f"Filtered regulations for {region_filter}: {[r['name'] for r in filtered_regulations]}")

        # Base query - Get all BOM entries for the SKU, joined with chemicals and regulations
        query = db.session.query(
            PFASBOM.component,
            PFASBOM.subcomponent,
            PFASBOM.material,
            PFASBOM.material_name,
            PFASMaterialChemicals.cas_number,
            PFASMaterialChemicals.chemical_name,
            PFASMaterialChemicals.concentration_ppm,
            PFASMaterialChemicals.supplier_name,
            PFASMaterialChemicals.reference_doc,
            PFASRegulations.australian_aics,
            PFASRegulations.australian_imap_tier_2,
            PFASRegulations.canadian_dsl,
            PFASRegulations.canada_pctsr_2012,
            PFASRegulations.eu_reach_pre_registered,
            PFASRegulations.eu_reach_registered_ppm,
            PFASRegulations.us_epa_tscainventory,
            PFASRegulations.us_epa_tsca12b
        ).outerjoin(
            PFASMaterialChemicals, PFASBOM.material == PFASMaterialChemicals.material_id
        ).outerjoin(
            PFASRegulations, PFASMaterialChemicals.cas_number == PFASRegulations.cas_number
        ).filter(PFASBOM.sku == sku)

        # Apply region filter if not 'all'
        # This means we only want to show rows where the selected regulation column has data
        if region_filter != 'all' and region_filter in region_to_columns:
            db_column_name = region_to_columns[region_filter]
            db_column = getattr(PFASRegulations, db_column_name)
            query = query.filter(db_column.isnot(None))
            app.logger.debug(f"Applied filter for column: {db_column_name}")

        results = query.all()
        app.logger.debug(f"Query returned {len(results)} rows after filtering.")

        data = []
        non_conforming_count = 0
        in_conformance_count = 0
        no_data_count = 0 # Includes entries with no chemical data OR entries where the specific regulation has no data

        for row in results:
            # Process each row
            component = row.component or "-"
            subcomponent = row.subcomponent or "-"
            material = row.material or "-"
            material_name = row.material_name or "-"
            cas = row.cas_number or "Unknown"
            chem_name = row.chemical_name or "Unknown"

            try:
                conc = int(float(row.concentration_ppm)) if row.concentration_ppm is not None else None
            except (ValueError, TypeError):
                conc = None

            supplier = row.supplier_name or "Unknown"
            has_chemical_data = row.cas_number is not None or row.chemical_name is not None

            limits = []
            is_non_conforming = False
            regulation_has_data = False # Flag to check if any of the *applied* regulations have data for this row

            # Loop through the *filtered* regulations list
            for reg_def in filtered_regulations:
                db_col_name = reg_def['col']
                value = getattr(row, db_col_name, None)
                threshold = float(value) if value is not None else None

                # Mark that at least one regulation column we are looking at has data
                if value is not None:
                    regulation_has_data = True

                if conc is None:
                    # 🚨 CRITICAL CHANGE: If concentration is unknown, mark as NON-CONFORMING for ALL applicable regulations
                    status = 'exceeded'
                    color = 'danger'
                    limit_display = f"{threshold} ppm" if threshold is not None else "No Threshold"
                    is_non_conforming = True
                else:
                    # Only calculate compliance if we have a valid concentration
                    if threshold is None:
                        status = 'unknown'
                        color = 'warning'
                        limit_display = "No Data"
                        # Don't set is_non_conforming here, as "No Data" means we don't know, not that it fails
                    elif threshold < conc:
                        status = 'exceeded'
                        color = 'danger'
                        is_non_conforming = True
                        limit_display = f"{threshold} ppm"
                    else:
                        status = 'within'
                        color = 'success'
                        limit_display = f"{threshold} ppm"

                limits.append({
                    'name': reg_def['name'],
                    'limit': limit_display,
                    'status': status,
                    'color': color
                })

            # Categorize entry based on results for the *filtered* regulations
            if not has_chemical_data:
                # No chemical data at all for this material entry
                no_data_count += 1
                entry_status = "No Chemical Data"
                status_color = "warning"
            elif not regulation_has_data and region_filter != 'all':
                 # Chemical data exists, but no data for the *specific* regulation we are filtering by
                 # Only count this if we are actually filtering by a region
                no_data_count += 1
                entry_status = "No Data for Region"
                status_color = "warning"
            elif is_non_conforming:
                # Failed one or more of the applicable regulations
                non_conforming_count += 1
                entry_status = "Non-Conforming"
                status_color = "danger"
            else:
                # Passed all applicable regulations (or no specific regulations were applied if 'all')
                in_conformance_count += 1
                entry_status = "Conforming"
                status_color = "success"

            data.append({
                "component": component,
                "subcomponent": subcomponent,
                "material": material,
                "material_name": material_name,
                "supplier_name": supplier,
                "chemical_name": chem_name,
                "cas_number": cas,
                "concentration": f"{int(conc)} ppm" if conc is not None else "Unknown",
                "reference_doc": row.reference_doc or "—",
                "limits": limits,
                "status": entry_status,
                "status_color": status_color
            })

        # Calculate summary statistics
        total_materials = db.session.query(PFASBOM.material).filter_by(sku=sku).distinct().count()
        materials_with_chemicals = db.session.query(PFASBOM.material).filter_by(sku=sku).join(
            PFASMaterialChemicals, PFASBOM.material == PFASMaterialChemicals.material_id
        ).distinct().count()

        summary = {
            'total': len(data),
            'non_conforming': non_conforming_count,
            'in_conformance': in_conformance_count,
            'no_chemical_data': no_data_count, # This now includes "No Data for Region"
            'files_total': total_materials,
            'files_downloaded': materials_with_chemicals,
            'files_not_found': max(0, total_materials - materials_with_chemicals),
            # These last two are derived from conformance counts, adjust logic if needed
            'alt_found': max(0, in_conformance_count // 3),
            'alt_not_found': max(0, non_conforming_count // 4)
        }

        return jsonify({
            'success': True,
            'region': region_filter,
            'data': data,
            'summary': summary,
            'regulations_applied': [reg['name'] for reg in filtered_regulations]
        })

    except Exception as e:
        app.logger.error(f"Error filtering assessment by region for SKU {sku}: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/assessment/<sku>')
def assessment_page(sku):
    """Unified assessment page with tab-based navigation using query params.
    
    Supports tabs: pfas, ppwr, rohs, reach
    URL format: /assessment/<sku>?tab=pfas&region=all
    """
    try:
        # Get actual route from database
        route_rec = db.session.query(Route).filter_by(sku=sku).first()
        if not route_rec:
            flash("SKU not found in system", "warning")
            return redirect(url_for('index'))
        
        actual_route = route_rec.route.lower()
        
        # Get requested tab (from query param or default to actual route)
        requested_tab = request.args.get('tab', actual_route).lower()
        
        # Get region parameter for filtering
        region_param = request.args.get('region', 'all')
        
        # Validate tab
        valid_tabs = ['pfas', 'ppwr', 'rohs', 'reach']
        if requested_tab not in valid_tabs:
            app.logger.warning(f"Invalid tab '{requested_tab}' requested, defaulting to {actual_route}")
            requested_tab = actual_route
        
        # Check if user is eligible for requested tab
        is_valid_tab = (requested_tab == actual_route)
        
        # Initialize data variables
        data = None
        summary_stats = None
        product_name = f"Product {sku}"
        
        # Load data ONLY if valid tab (prevents unauthorized access to wrong assessment type)
        if is_valid_tab:
            if requested_tab == 'pfas':
                # Load PFAS data from ppwr_bom + result tables
                app.logger.info(f"Loading PFAS assessment for SKU {sku}")
                
                # Helper function to normalize response types
                def _extract_json_from_response(resp):
                    if isinstance(resp, dict):
                        return resp, 200
                    if isinstance(resp, tuple) and len(resp) >= 1:
                        body = resp[0]
                        status = resp[1] if len(resp) > 1 and isinstance(resp[1], int) else None
                        if hasattr(body, 'get_json'):
                            try:
                                return body.get_json(), status or getattr(body, 'status_code', 200)
                            except Exception:
                                return {"error": "Invalid JSON in internal response"}, status or 500
                        elif isinstance(body, dict):
                            return body, status or 200
                        else:
                            return {"error": "Unexpected response type from internal call"}, status or 500
                    if hasattr(resp, 'get_json') and hasattr(resp, 'status_code'):
                        try:
                            return resp.get_json(), resp.status_code
                        except Exception:
                            return {"error": "Invalid JSON in internal response"}, resp.status_code
                    return {"error": "Unknown response type from internal call"}, 500
                
                # Get PFAS assessment data (with optional region filtering)
                if region_param != 'all':
                    filter_resp = filter_assessment_by_region(sku)
                    filter_data, filter_status = _extract_json_from_response(filter_resp)
                    if filter_status == 200 and filter_data.get('success', False):
                        data = {
                            'product': f"{sku}_Filtered",
                            'summary': filter_data.get('summary', {}),
                            'data': filter_data.get('data', [])
                        }
                    else:
                        resp = pfas_assessment(sku)
                        data, status = _extract_json_from_response(resp)
                else:
                    resp = pfas_assessment(sku)
                    data, status = _extract_json_from_response(resp)
                
                if data and "error" not in data:
                    summary_stats = calculate_dynamic_summary(sku, data, strict=False)
                    product_name = data.get("product", f"{sku}_PFAS")
                
            elif requested_tab == 'ppwr':
                # Load PPWR data from ppwr_bom + ppwr_result tables (unified RAG + manual data)
                app.logger.info(f"Loading PPWR assessment for SKU {sku}")
                
                # Re-use PFAS data structure but apply strict rules
                resp = pfas_assessment(sku)
                
                def _extract(resp_obj):
                    if isinstance(resp_obj, dict):
                        return resp_obj, 200
                    if isinstance(resp_obj, tuple) and len(resp_obj) >= 1:
                        body = resp_obj[0]
                        status = resp_obj[1] if len(resp_obj) > 1 and isinstance(resp_obj[1], int) else 200
                        if hasattr(body, 'get_json'):
                            try:
                                return body.get_json(), status
                            except Exception:
                                return {"error": "Invalid JSON in internal response"}, 500
                        elif isinstance(body, dict):
                            return body, status
                        else:
                            return {"error": "Unexpected internal response type"}, 500
                    if hasattr(resp_obj, 'get_json') and hasattr(resp_obj, 'status_code'):
                        try:
                            return resp_obj.get_json(), resp_obj.status_code
                        except Exception:
                            return {"error": "Invalid JSON in internal response"}, 500
                    return {"error": "Unknown response type"}, 500
                
                data, status = _extract(resp)
                
                if data and "error" not in data and status == 200:
                    # Post-process: enforce strict PPWR rules (unknown concentration = non-conforming)
                    for entry in data.get('data', []):
                        conc_raw = entry.get('concentration', '')
                        if not conc_raw or str(conc_raw).strip().lower() == 'unknown':
                            entry['status'] = 'Non-Conforming'
                            entry['status_color'] = 'danger'
                    
                    summary_stats = calculate_dynamic_summary(sku, data, strict=True)
                    product_name = data.get('product', f"{sku}_PPWR")
            
            elif requested_tab == 'rohs':
                # TODO: Implement RoHS data loading from rohs_bom + rohs_assessments tables
                app.logger.info(f"RoHS assessment not yet implemented for SKU {sku}")
                data = {"data": [], "product": f"{sku}_RoHS"}
                summary_stats = {}
                product_name = f"{sku}_RoHS"
            
            elif requested_tab == 'reach':
                # TODO: Implement REACH data loading
                app.logger.info(f"REACH assessment not yet implemented for SKU {sku}")
                data = {"data": [], "product": f"{sku}_REACH"}
                summary_stats = {}
                product_name = f"{sku}_REACH"
        
        # Always render template (even for invalid tabs - they'll show eligibility message)
        return render_template('assessment.html',
                             sku=sku,
                             product_name=product_name,
                             active_tab=requested_tab,
                             actual_route=actual_route,
                             is_valid_tab=is_valid_tab,
                             initial_data=data,
                             summary_stats=summary_stats,
                             selected_region=region_param,
                             available_tabs=valid_tabs)

    except Exception as e:
        app.logger.error(f"Unhandled error in assessment_page for {sku}: {e}", exc_info=True)
        flash("Failed to load assessment. Please try again later.", "danger")
        return redirect(url_for('index'))


@app.route('/static/templates/ppwr_bom_template.xlsx')
def download_ppwr_template():
    """Serve the PPWR BOM template from Helper_Data as a downloadable file.
    This avoids duplicating the file into the frontend static folder.
    """
    try:
        helper_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'Helper_Data'))
        filename = 'BOM-1_VET SYRINGE (3).xlsx'
        file_path = os.path.join(helper_dir, filename)
        if not os.path.exists(file_path):
            app.logger.error(f"PPWR template not found at {file_path}")
            return jsonify({'error': 'Template not found'}), 404

        return send_file(
            file_path,
            as_attachment=True,
            download_name='ppwr_bom_template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        app.logger.error(f"Error serving PPWR template: {e}", exc_info=True)
        return jsonify({'error': 'Failed to serve template'}), 500


@app.route('/ppwr-assessment/<sku>')
def ppwr_assessment_page(sku):
    """Legacy route: Redirects to unified assessment page with tab=ppwr.
    
    This maintains backward compatibility while migrating to the new unified tab structure.
    """
    try:
        # Preserve region parameter if provided
        region_param = request.args.get('region', 'all')
        
        # Redirect to unified route with ppwr tab
        return redirect(url_for('assessment_page', sku=sku, tab='ppwr', region=region_param), code=301)
        
    except Exception as e:
        app.logger.error(f"PPWR redirect failed for {sku}: {e}", exc_info=True)
        return redirect(url_for('assessment_page', sku=sku, tab='ppwr'))

    except Exception as e:
        app.logger.error(f"Unhandled error in ppwr_assessment_page for {sku}: {e}", exc_info=True)
        flash("Failed to load PPWR assessment. Please try again later.", "danger")
        return redirect(url_for('index'))

@app.errorhandler(404)
def not_found(error):
    return jsonify({'error': 'Endpoint not found'}), 404

@app.errorhandler(500)
def internal_error(error):
    return jsonify({'error': 'Internal server error'}), 500

@app.route('/logs')
def view_logs():
    """Render a simple log viewer for the frontend app logs.

    Reads the last N lines from frontend/logs/app.log and displays them with basic formatting.
    """
    try:
        logs_dir = os.path.join(os.path.dirname(__file__), 'logs')
        log_path = os.path.join(logs_dir, 'app.log')
        lines = []
        if os.path.exists(log_path):
            with open(log_path, 'r', encoding='utf-8', errors='replace') as f:
                lines = f.readlines()
        # Show last 400 lines to keep page light
        tail = lines[-400:] if lines else []
        return render_template('logs.html', log_lines=tail)
    except Exception as e:
        app.logger.error(f"Failed to render logs: {e}", exc_info=True)
        flash('Unable to load logs', 'danger')
        return render_template('logs.html', log_lines=[])

if __name__ == '__main__':
    # Robust startup: retry DB initialization a few times to avoid noisy errors
    # when the database container isn't ready yet.
    max_attempts = 10
    delay_seconds = 2

    with app.app_context():
        for attempt in range(1, max_attempts + 1):
            try:
                db.create_all()
                # Ensure the pfas_bom table has the uploaded_at column (safe idempotent alter)
                db.session.execute(text("ALTER TABLE pfas_bom ADD COLUMN IF NOT EXISTS uploaded_at TIMESTAMP"))
                db.session.commit()
                app.logger.info("Database tables created and uploaded_at ensured (attempt %d)", attempt)
                break
            except Exception as e:
                db.session.rollback()
                app.logger.warning(f"Database init attempt {attempt} failed: {e}")
                if attempt == max_attempts:
                    app.logger.error("Giving up on DB init after %d attempts; app will still start.", max_attempts)
                else:
                    import time
                    time.sleep(delay_seconds)

    # Allow overriding port via PORT env var for local runs
    try:
        run_port = int(os.environ.get('PORT', '5000'))
    except Exception:
        run_port = 5000
    # Bind host: in Docker expose on 0.0.0.0, otherwise loopback
    try:
        running_in_docker = os.path.exists('/.dockerenv') or os.environ.get('DOCKER') == '1'
    except Exception:
        running_in_docker = False
    bind_host = '0.0.0.0' if running_in_docker else '127.0.0.1'
    app.run(debug=True, host=bind_host, port=run_port, use_reloader=False)